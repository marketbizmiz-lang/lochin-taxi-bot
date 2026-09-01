import os
import re
import io
import time
import json
import random
import asyncio
import logging
import sqlite3
import base64
import hashlib
from pathlib import Path
from typing import Any, Optional, List, Set, Tuple, Dict, Callable, Awaitable
from datetime import datetime, timezone, timedelta

import aiohttp
import asyncpg
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from aiohttp import web

# Cryptography xavfsiz modul tekshiruvi (Hech qachon ModuleNotFoundError bermaydi)
try:
    from cryptography.fernet import Fernet
    HAS_CRYPTOGRAPHY = True
except ImportError:
    HAS_CRYPTOGRAPHY = False

from aiogram import Bot, Dispatcher, F, Router, BaseMiddleware
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command, CommandStart, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    Message,
    CallbackQuery,
    KeyboardButton,
    ReplyKeyboardMarkup,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    BufferedInputFile,
    TelegramObject,
)

# ============================================================
# 1. ASOSIY SOZLAMALAR VA AVTOMATIK XAVFSIZLIK
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
logger = logging.getLogger("lochin_taxi_bot")

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "lochin_taxi.db"
AUDIT_LOG_PATH = BASE_DIR / "audit.log"
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()

BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
BOT_NAME = os.getenv("BOT_NAME", "LOCHIN TAXI").strip() or "LOCHIN TAXI"
PORT = int(os.getenv("PORT", "8080"))

# ADMIN_IDS — .env orqali olinadi
ADMIN_IDS: Set[int] = set()
_env_admins = os.getenv("ADMIN_IDS", "").strip()
if _env_admins:
    for _adm in _env_admins.split(","):
        _adm_clean = _adm.strip()
        if _adm_clean.isdigit():
            ADMIN_IDS.add(int(_adm_clean))

MANAGER_TG_ID = int(os.getenv("MANAGER_TG_ID", "0"))
SUPPORT_PHONE = os.getenv("SUPPORT_PHONE", "+998913773200").strip()
SUPPORT_PHONE_DISPLAY = os.getenv("SUPPORT_PHONE_DISPLAY", "+998 91 377 32 00").strip()
DRIVER_GROUP_LINK = os.getenv("DRIVER_GROUP_LINK", "https://t.me/+vLyCiiXNvB5kMTUy").strip()

# ENCRYPTION_KEY — Avtomatik yaratiladi, hech qachon RuntimeError bermaydi!
raw_key = os.getenv("ENCRYPTION_KEY", "").strip()
if not raw_key or len(raw_key) < 16:
    raw_key = hashlib.sha256((BOT_TOKEN or "LOCHIN_TAXI_DEFAULT_SALT_2026").encode()).hexdigest()

ENCRYPTION_KEY = raw_key
_cipher_suite = None

if HAS_CRYPTOGRAPHY:
    try:
        derived_key = base64.urlsafe_b64encode(hashlib.sha256(ENCRYPTION_KEY.encode()).digest())
        _cipher_suite = Fernet(derived_key)
    except Exception:
        pass

if not _cipher_suite:
    class PureCipher:
        def __init__(self, key: str):
            self.k = hashlib.sha256(key.encode()).digest()
        def encrypt(self, data: bytes) -> bytes:
            rep = (self.k * (len(data) // len(self.k) + 1))[:len(data)]
            return base64.urlsafe_b64encode(bytes(a ^ b for a, b in zip(data, rep)))
        def decrypt(self, data: bytes) -> bytes:
            raw = base64.urlsafe_b64decode(data)
            rep = (self.k * (len(raw) // len(self.k) + 1))[:len(raw)]
            return bytes(a ^ b for a, b in zip(raw, rep))
    _cipher_suite = PureCipher(ENCRYPTION_KEY)

# Yandex Fleet API
YANDEX_API_KEY = os.getenv("YANDEX_API_KEY", "").strip()
YANDEX_CLIENT_ID = os.getenv("YANDEX_CLIENT_ID", "").strip()
YANDEX_PARK_ID = os.getenv("YANDEX_PARK_ID", "").strip()
YANDEX_FLEET_URL = "https://fleet-api.taxi.yandex.net"

# Moliyaviy parametrlar (Faqat butun so'm — INTEGER)
MIN_WITHDRAWAL = int(os.getenv("MIN_WITHDRAWAL", "20000"))
MIN_DEPOSIT = int(os.getenv("MIN_DEPOSIT", "20000"))
COMMISSION_PERCENT = float(os.getenv("COMMISSION_PERCENT", "0.0"))

TASHKENT_TZ = timezone(timedelta(hours=5))

UZ_MONTHS = {
    1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May", 6: "Iyun",
    7: "Iyul", 8: "Avgust", 9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr"
}


# ============================================================
# 2. XAVFSIZLIK VA YORDAMCHI FUNKSIYALAR
# ============================================================

def encrypt_card(card_number: str) -> str:
    """Karta raqamini bazaga saqlashdan oldin shifrlaydi."""
    if not card_number:
        return ""
    clean = re.sub(r"\D", "", str(card_number))
    try:
        return _cipher_suite.encrypt(clean.encode()).decode()
    except Exception:
        return clean


def decrypt_card(encrypted_card: str) -> str:
    """Bazadagi shifrlangan kartani asl 16 talik raqamiga qaytaradi."""
    if not encrypted_card:
        return ""
    try:
        return _cipher_suite.decrypt(encrypted_card.encode()).decode()
    except Exception:
        return encrypted_card


def mask_card(card_number: str) -> str:
    """Kartani 8600 **** **** 1234 ko'rinishida xavfsiz maskalaydi."""
    clean = re.sub(r"\D", "", str(card_number))
    if len(clean) == 16:
        return f"{clean[:4]} **** **** {clean[-4:]}"
    elif len(clean) >= 8:
        return f"{clean[:4]} **** {clean[-4:]}"
    return "Noma'lum karta"


def log_admin_view_card(admin_id: int, withdrawal_id: int):
    """Admin to'liq kartani ko'rganda audit log fayliga yozadi."""
    iso_time = tashkent_now_iso()
    logger.info(f"AUDIT | ADMIN {admin_id} viewed full card for withdrawal_id={withdrawal_id}")
    try:
        with open(AUDIT_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(f"{iso_time} | ADMIN_ID:{admin_id} | WITHDRAWAL_ID:{withdrawal_id}\n")
    except Exception as e:
        logger.error(f"Audit log yozishda xatolik: {e}")


def fmt_sum(val: Any) -> str:
    """Faqat butun so'm formatida ajratib ko'rsatish."""
    try:
        if val is None:
            return "0"
        return f"{int(val):,}".replace(",", " ")
    except Exception:
        return "0"


def clean_phone_number(raw_phone: str) -> str:
    """Telefon raqamni qat'iy +998XXXXXXXXX formatga keltiradi."""
    digits = re.sub(r"\D", "", str(raw_phone or ""))
    if digits.startswith("8") and len(digits) == 11:
        digits = "998" + digits[1:]
    elif not digits.startswith("998") and len(digits) == 9:
        digits = "998" + digits
    elif digits.startswith("998") and len(digits) == 12:
        pass
    else:
        if len(digits) >= 9:
            digits = "998" + digits[-9:]
    return f"+{digits}"


def tashkent_now_iso() -> str:
    return datetime.now(TASHKENT_TZ).replace(microsecond=0).isoformat()


# ============================================================
# 3. DATABASE LAYER (FOR UPDATE LOCKING & ATOMIK TRANSACTIONS)
# ============================================================

db_pool: Optional[asyncpg.Pool] = None


async def init_database():
    global db_pool
    if DATABASE_URL:
        try:
            clean_url = DATABASE_URL.replace("?sslmode=require", "")
            db_pool = await asyncpg.create_pool(
                clean_url, ssl="require", min_size=5, max_size=30, timeout=15
            )
            async with db_pool.acquire() as conn:
                await conn.execute("""
                    CREATE TABLE IF NOT EXISTS users (
                        id                SERIAL PRIMARY KEY,
                        telegram_id       BIGINT UNIQUE NOT NULL,
                        username          TEXT,
                        full_name         TEXT,
                        phone             TEXT UNIQUE,
                        card_number       TEXT,
                        car_model         TEXT,
                        car_number        TEXT,
                        position          TEXT UNIQUE,
                        language          TEXT NOT NULL DEFAULT 'uz',
                        balance           BIGINT DEFAULT 0,
                        blocked_balance   BIGINT DEFAULT 0,
                        is_registered     INT DEFAULT 0,
                        is_blocked        INT DEFAULT 0,
                        yandex_driver_id  TEXT UNIQUE,
                        referrer_id       BIGINT,
                        total_orders      INT DEFAULT 0,
                        total_earnings    BIGINT DEFAULT 0,
                        last_activity     TEXT,
                        created_at        TEXT NOT NULL,
                        updated_at        TEXT NOT NULL
                    );
                    CREATE INDEX IF NOT EXISTS idx_users_phone ON users(phone);
                    CREATE INDEX IF NOT EXISTS idx_users_tg_id ON users(telegram_id);
                    CREATE INDEX IF NOT EXISTS idx_users_pos ON users(position);
                    CREATE INDEX IF NOT EXISTS idx_users_y_id ON users(yandex_driver_id);

                    CREATE TABLE IF NOT EXISTS withdrawals (
                        id             SERIAL PRIMARY KEY,
                        user_id        INT NOT NULL,
                        amount         BIGINT NOT NULL,
                        commission     BIGINT DEFAULT 0,
                        net_amount     BIGINT NOT NULL,
                        card_number    TEXT,
                        status         TEXT NOT NULL DEFAULT 'pending',
                        payout_method  TEXT DEFAULT 'manual',
                        ext_tx_id      TEXT,
                        created_at     TEXT NOT NULL,
                        updated_at     TEXT NOT NULL
                    );
                    CREATE INDEX IF NOT EXISTS idx_wd_user_id ON withdrawals(user_id);
                    CREATE INDEX IF NOT EXISTS idx_wd_status ON withdrawals(status);
                    CREATE INDEX IF NOT EXISTS idx_wd_created ON withdrawals(created_at);
                """)
            logger.info("PostgreSQL (asyncpg) muvaffaqiyatli ishga tushdi!")
        except Exception as e:
            logger.error(f"PostgreSQL ulanishida xatolik: {e}. SQLite WAL rejimiga o'tilmoqda.")
            db_pool = None

    if not db_pool:
        conn = sqlite3.connect(DB_PATH, timeout=20)
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        conn.execute("PRAGMA busy_timeout=5000;")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_id      INTEGER UNIQUE NOT NULL,
                username         TEXT,
                full_name        TEXT,
                phone            TEXT UNIQUE,
                card_number      TEXT,
                car_model        TEXT,
                car_number       TEXT,
                position         TEXT UNIQUE,
                language         TEXT NOT NULL DEFAULT 'uz',
                balance          INTEGER DEFAULT 0,
                blocked_balance  INTEGER DEFAULT 0,
                is_registered    INTEGER DEFAULT 0,
                is_blocked       INTEGER DEFAULT 0,
                yandex_driver_id TEXT UNIQUE,
                referrer_id      INTEGER,
                total_orders     INTEGER DEFAULT 0,
                total_earnings   INTEGER DEFAULT 0,
                last_activity    TEXT,
                created_at       TEXT NOT NULL,
                updated_at       TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS withdrawals (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id       INTEGER NOT NULL,
                amount        INTEGER NOT NULL,
                commission    INTEGER DEFAULT 0,
                net_amount    INTEGER NOT NULL,
                card_number   TEXT,
                status        TEXT NOT NULL DEFAULT 'pending',
                payout_method TEXT DEFAULT 'manual',
                ext_tx_id     TEXT,
                created_at    TEXT NOT NULL,
                updated_at    TEXT NOT NULL
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_users_phone ON users(phone);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_users_pos ON users(position);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_users_y_id ON users(yandex_driver_id);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_wd_user ON withdrawals(user_id);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_wd_status ON withdrawals(status);")
        conn.commit()
        conn.close()
        logger.info("SQLite (WAL High Speed Mode) tayyor!")


def _process_user_dict(d: Optional[dict]) -> Optional[dict]:
    if not d:
        return None
    res = dict(d)
    if "card_number" in res and res["card_number"]:
        res["card_number"] = decrypt_card(res["card_number"])
    if "balance" in res:
        res["balance"] = int(res["balance"] or 0)
    return res


def _process_wd_dict(d: Optional[dict]) -> Optional[dict]:
    if not d:
        return None
    res = dict(d)
    if "card_number" in res and res["card_number"]:
        res["card_number"] = decrypt_card(res["card_number"])
    if "amount" in res:
        res["amount"] = int(res["amount"] or 0)
    if "net_amount" in res:
        res["net_amount"] = int(res["net_amount"] or 0)
    if "commission" in res:
        res["commission"] = int(res["commission"] or 0)
    return res


async def db_get_user(telegram_id: int) -> Optional[dict]:
    if db_pool:
        async with db_pool.acquire() as conn:
            row = await conn.fetchrow("SELECT * FROM users WHERE telegram_id = $1", telegram_id)
            return _process_user_dict(dict(row)) if row else None
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM users WHERE telegram_id = ?", (telegram_id,)).fetchone()
        conn.close()
        return _process_user_dict(dict(row)) if row else None


async def db_get_user_by_id(user_id: int) -> Optional[dict]:
    if db_pool:
        async with db_pool.acquire() as conn:
            row = await conn.fetchrow("SELECT * FROM users WHERE id = $1", user_id)
            return _process_user_dict(dict(row)) if row else None
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        conn.close()
        return _process_user_dict(dict(row)) if row else None


async def db_get_user_by_phone(phone: str) -> Optional[dict]:
    clean_p = clean_phone_number(phone)
    if db_pool:
        async with db_pool.acquire() as conn:
            row = await conn.fetchrow("SELECT * FROM users WHERE phone = $1", clean_p)
            return _process_user_dict(dict(row)) if row else None
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM users WHERE phone = ?", (clean_p,)).fetchone()
        conn.close()
        return _process_user_dict(dict(row)) if row else None


async def db_find_driver_by_query(query: str) -> Optional[dict]:
    clean_q = query.strip()
    phone_clean = clean_phone_number(clean_q)
    if db_pool:
        async with db_pool.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT * FROM users WHERE position ILIKE $1 OR phone = $2 OR car_number ILIKE $1",
                f"%{clean_q}%", phone_clean
            )
            return _process_user_dict(dict(row)) if row else None
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT * FROM users WHERE position LIKE ? OR phone = ? OR car_number LIKE ?",
            (f"%{clean_q}%", phone_clean, f"%{clean_q}%")
        ).fetchone()
        conn.close()
        return _process_user_dict(dict(row)) if row else None


async def db_delete_user_by_id(user_id: int) -> bool:
    try:
        if db_pool:
            async with db_pool.acquire() as conn:
                async with conn.transaction():
                    await conn.execute("DELETE FROM withdrawals WHERE user_id = $1", user_id)
                    await conn.execute("DELETE FROM users WHERE id = $1", user_id)
                return True
        else:
            conn = sqlite3.connect(DB_PATH, timeout=10)
            with conn:
                conn.execute("DELETE FROM withdrawals WHERE user_id = ?", (user_id,))
                conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
            conn.close()
            return True
    except Exception as e:
        logger.error(f"db_delete_user_by_id xatosi: {e}")
        return False


async def db_upsert_start(telegram_id: int, username: str):
    now = tashkent_now_iso()
    if db_pool:
        async with db_pool.acquire() as conn:
            await conn.execute(
                """
                INSERT INTO users (telegram_id, username, last_activity, created_at, updated_at)
                VALUES ($1, $2, $3, $4, $5)
                ON CONFLICT (telegram_id) DO UPDATE SET last_activity = $3, updated_at = $5
                """,
                telegram_id, username or "", now, now, now,
            )
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        with conn:
            conn.execute(
                """
                INSERT INTO users (telegram_id, username, last_activity, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(telegram_id) DO UPDATE SET last_activity=excluded.last_activity, updated_at=excluded.updated_at
                """,
                (telegram_id, username or "", now, now, now),
            )
        conn.close()


async def db_set_language(telegram_id: int, language: str):
    now = tashkent_now_iso()
    if db_pool:
        async with db_pool.acquire() as conn:
            await conn.execute(
                "UPDATE users SET language = $1, updated_at = $2 WHERE telegram_id = $3",
                language, now, telegram_id,
            )
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        with conn:
            conn.execute(
                "UPDATE users SET language = ?, updated_at = ? WHERE telegram_id = ?",
                (language, now, telegram_id),
            )
        conn.close()


async def db_generate_unique_position() -> str:
    for _ in range(200):
        pos = f"LCH-{random.randint(1000, 9999)}"
        if db_pool:
            async with db_pool.acquire() as conn:
                exists = await conn.fetchval("SELECT 1 FROM users WHERE position = $1", pos)
                if not exists:
                    return pos
        else:
            conn = sqlite3.connect(DB_PATH, timeout=10)
            row = conn.execute("SELECT 1 FROM users WHERE position = ?", (pos,)).fetchone()
            conn.close()
            if not row:
                return pos
    return f"LCH-{random.randint(10000, 99999)}"


async def db_finish_registration(
    telegram_id: int, full_name: str, phone: str, card_number: str,
    car_model: str, car_number: str, yandex_driver_id: Optional[str],
) -> str:
    position = await db_generate_unique_position()
    now = tashkent_now_iso()
    enc_card = encrypt_card(card_number)
    phone_clean = clean_phone_number(phone)

    if db_pool:
        async with db_pool.acquire() as conn:
            await conn.execute(
                """UPDATE users SET 
                    full_name=$1, phone=$2, card_number=$3, car_model=$4, 
                    car_number=$5, position=$6, yandex_driver_id=$7, is_registered=1, 
                    last_activity=$8, updated_at=$8 
                WHERE telegram_id=$9""",
                full_name, phone_clean, enc_card, car_model, car_number, position, yandex_driver_id, now, telegram_id,
            )
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        with conn:
            conn.execute(
                """UPDATE users SET 
                    full_name=?, phone=?, card_number=?, car_model=?, 
                    car_number=?, position=?, yandex_driver_id=?, is_registered=1, 
                    last_activity=?, updated_at=? 
                WHERE telegram_id=?""",
                (full_name, phone_clean, enc_card, car_model, car_number, position, yandex_driver_id, now, now, telegram_id),
            )
        conn.close()
    return position


async def db_update_balance(telegram_id: int, balance: int):
    now = tashkent_now_iso()
    int_bal = int(balance)
    if db_pool:
        async with db_pool.acquire() as conn:
            await conn.execute(
                "UPDATE users SET balance=$1, updated_at=$2 WHERE telegram_id=$3",
                int_bal, now, telegram_id,
            )
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        with conn:
            conn.execute(
                "UPDATE users SET balance=?, updated_at=? WHERE telegram_id=?",
                (int_bal, now, telegram_id),
            )
        conn.close()


async def db_create_withdrawal(
    user_id: int, telegram_id: int, amount: int, commission: int,
    net_amount: int, card_number: str, status: str, payout_method: str, ext_tx_id: str = "",
) -> int:
    now = tashkent_now_iso()
    enc_card = encrypt_card(card_number)

    if db_pool:
        async with db_pool.acquire() as conn:
            async with conn.transaction():
                pending = await conn.fetchval(
                    "SELECT 1 FROM withdrawals WHERE user_id = $1 AND status = 'pending' FOR UPDATE", user_id
                )
                if pending:
                    raise ValueError("Sizda allaqachon ko'rib chiqilayotgan faol ariza mavjud!")

                cur_bal = await conn.fetchval(
                    "SELECT balance FROM users WHERE id = $1 FOR UPDATE", user_id
                )
                if cur_bal is None:
                    raise ValueError("Foydalanuvchi topilmadi")
                
                avail = cur_bal - MIN_DEPOSIT
                if avail < amount:
                    raise ValueError("Yetarli mablag' mavjud emas yoki balans o'zgargan")

                await conn.execute(
                    "UPDATE users SET balance = balance - $1, updated_at = $2 WHERE id = $3",
                    amount, now, user_id,
                )
                row = await conn.fetchrow(
                    """INSERT INTO withdrawals 
                        (user_id, amount, commission, net_amount, card_number, status, payout_method, ext_tx_id, created_at, updated_at) 
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10) RETURNING id""",
                    user_id, amount, commission, net_amount, enc_card, status, payout_method, ext_tx_id, now, now,
                )
                return row["id"] if row else 0
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        with conn:
            cur = conn.cursor()
            cur.execute("SELECT 1 FROM withdrawals WHERE user_id = ? AND status = 'pending'", (user_id,))
            if cur.fetchone():
                raise ValueError("Sizda allaqachon ko'rib chiqilayotgan faol ariza mavjud!")

            cur.execute("SELECT balance FROM users WHERE id = ?", (user_id,))
            row = cur.fetchone()
            if not row:
                raise ValueError("Foydalanuvchi topilmadi")
            
            cur_bal = int(row[0])
            avail = cur_bal - MIN_DEPOSIT
            if avail < amount:
                raise ValueError("Yetarli mablag' mavjud emas yoki balans o'zgargan")

            cur.execute(
                "UPDATE users SET balance = balance - ?, updated_at = ? WHERE id = ?",
                (amount, now, user_id),
            )
            cur.execute(
                """INSERT INTO withdrawals 
                    (user_id, amount, commission, net_amount, card_number, status, payout_method, ext_tx_id, created_at, updated_at) 
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (user_id, amount, commission, net_amount, enc_card, status, payout_method, ext_tx_id, now, now),
            )
            w_id = cur.lastrowid
        conn.close()
        return w_id or 0


async def db_get_withdrawal(w_id: int) -> Optional[dict]:
    if db_pool:
        async with db_pool.acquire() as conn:
            row = await conn.fetchrow("SELECT * FROM withdrawals WHERE id = $1", w_id)
            return _process_wd_dict(dict(row)) if row else None
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM withdrawals WHERE id = ?", (w_id,)).fetchone()
        conn.close()
        return _process_wd_dict(dict(row)) if row else None


async def db_update_withdrawal_status(w_id: int, status: str, ext_tx_id: str = ""):
    now = tashkent_now_iso()
    if db_pool:
        async with db_pool.acquire() as conn:
            await conn.execute(
                "UPDATE withdrawals SET status=$1, ext_tx_id=$2, updated_at=$3 WHERE id=$4",
                status, ext_tx_id, now, w_id,
            )
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        with conn:
            conn.execute(
                "UPDATE withdrawals SET status=?, ext_tx_id=?, updated_at=? WHERE id=?",
                (status, ext_tx_id, now, w_id),
            )
        conn.close()


async def db_refund_withdrawal(w_id: int):
    wd = await db_get_withdrawal(w_id)
    if not wd or wd.get("status") in ("rejected", "completed"):
        return
    now = tashkent_now_iso()
    amount = int(wd["amount"])
    user_id = wd["user_id"]

    if db_pool:
        async with db_pool.acquire() as conn:
            async with conn.transaction():
                await conn.execute("UPDATE users SET balance = balance + $1, updated_at = $2 WHERE id = $3", amount, now, user_id)
                await conn.execute("UPDATE withdrawals SET status='rejected', updated_at=$1 WHERE id=$2", now, w_id)
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        with conn:
            conn.execute("UPDATE users SET balance = balance + ?, updated_at = ? WHERE id = ?", (amount, now, user_id))
            conn.execute("UPDATE withdrawals SET status='rejected', updated_at=? WHERE id=?", (now, w_id))
        conn.close()


async def db_get_driver_today_withdrawn(user_id: int) -> int:
    now_tashkent = datetime.now(TASHKENT_TZ)
    today_start = now_tashkent.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    if db_pool:
        async with db_pool.acquire() as conn:
            val = await conn.fetchval(
                "SELECT COALESCE(SUM(amount),0) FROM withdrawals WHERE user_id = $1 AND status='completed' AND updated_at >= $2",
                user_id, today_start
            )
            return int(val or 0)
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        row = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM withdrawals WHERE user_id = ? AND status='completed' AND updated_at >= ?",
            (user_id, today_start)
        ).fetchone()
        conn.close()
        return int(row[0] if row else 0)


async def db_get_all_registered_drivers() -> List[dict]:
    if db_pool:
        async with db_pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM users WHERE is_registered=1 ORDER BY id ASC")
            return [_process_user_dict(dict(r)) for r in rows]
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT * FROM users WHERE is_registered=1 ORDER BY id ASC").fetchall()
        conn.close()
        return [_process_user_dict(dict(r)) for r in rows]


async def db_get_all_users() -> List[dict]:
    if db_pool:
        async with db_pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM users ORDER BY id ASC")
            return [_process_user_dict(dict(r)) for r in rows]
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT * FROM users ORDER BY id ASC").fetchall()
        conn.close()
        return [_process_user_dict(dict(r)) for r in rows]


async def db_get_stats() -> dict:
    now_tashkent = datetime.now(TASHKENT_TZ)
    today_start = now_tashkent.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    month_start = now_tashkent.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()

    if db_pool:
        async with db_pool.acquire() as conn:
            total_users     = await conn.fetchval("SELECT COUNT(*) FROM users") or 0
            registered      = await conn.fetchval("SELECT COUNT(*) FROM users WHERE is_registered=1") or 0
            yandex_linked   = await conn.fetchval("SELECT COUNT(*) FROM users WHERE yandex_driver_id IS NOT NULL AND yandex_driver_id!=''") or 0
            
            today_withdrawn = await conn.fetchval("SELECT COALESCE(SUM(amount),0) FROM withdrawals WHERE status='completed' AND updated_at >= $1", today_start) or 0
            month_withdrawn = await conn.fetchval("SELECT COALESCE(SUM(amount),0) FROM withdrawals WHERE status='completed' AND updated_at >= $1", month_start) or 0
            total_withdrawn = await conn.fetchval("SELECT COALESCE(SUM(amount),0) FROM withdrawals WHERE status='completed'") or 0
            
            pending_count   = await conn.fetchval("SELECT COUNT(*) FROM withdrawals WHERE status='pending'") or 0
            pending_sum     = await conn.fetchval("SELECT COALESCE(SUM(amount),0) FROM withdrawals WHERE status='pending'") or 0
            total_comm      = await conn.fetchval("SELECT COALESCE(SUM(commission),0) FROM withdrawals WHERE status='completed'") or 0
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        total_users     = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        registered      = conn.execute("SELECT COUNT(*) FROM users WHERE is_registered=1").fetchone()[0]
        yandex_linked   = conn.execute("SELECT COUNT(*) FROM users WHERE yandex_driver_id IS NOT NULL AND yandex_driver_id!=''").fetchone()[0]
        
        today_withdrawn = conn.execute("SELECT COALESCE(SUM(amount),0) FROM withdrawals WHERE status='completed' AND updated_at >= ?", (today_start,)).fetchone()[0]
        month_withdrawn = conn.execute("SELECT COALESCE(SUM(amount),0) FROM withdrawals WHERE status='completed' AND updated_at >= ?", (month_start,)).fetchone()[0]
        total_withdrawn = conn.execute("SELECT COALESCE(SUM(amount),0) FROM withdrawals WHERE status='completed'").fetchone()[0]
        
        pending_count   = conn.execute("SELECT COUNT(*) FROM withdrawals WHERE status='pending'").fetchone()[0]
        pending_sum     = conn.execute("SELECT COALESCE(SUM(amount),0) FROM withdrawals WHERE status='pending'").fetchone()[0]
        total_comm      = conn.execute("SELECT COALESCE(SUM(commission),0) FROM withdrawals WHERE status='completed'").fetchone()[0]
        conn.close()

    return {
        "total_users": total_users,
        "registered_drivers": registered,
        "yandex_linked": yandex_linked,
        "today_withdrawn": int(today_withdrawn),
        "month_withdrawn": int(month_withdrawn),
        "total_withdrawn": int(total_withdrawn),
        "pending_count": pending_count,
        "pending_sum": int(pending_sum),
        "total_comm": int(total_comm),
    }


# ============================================================
# 4. YANDEX FLEET API
# ============================================================

class YandexFleetAPI:
    FLEET_BASE = "https://fleet-api.taxi.yandex.net"

    def __init__(self, api_key: str, client_id: str, park_id: str):
        self.api_key = api_key.strip()
        self.client_id = client_id.strip()
        self.park_id = park_id.strip()
        self._session: Optional[aiohttp.ClientSession] = None
        self._drivers_cache: List[dict] = []
        self._cache_ts: Optional[datetime] = None
        self._cache_ttl = 180
        self._balance_cache: Dict[str, Tuple[int, datetime]] = {}

    def _is_configured(self) -> bool:
        return bool(self.api_key and self.park_id and self.client_id)

    @property
    def _headers(self) -> dict:
        return {
            "X-Client-ID": self.client_id,
            "X-API-Key": self.api_key,
            "X-Park-ID": self.park_id,
            "Content-Type": "application/json",
            "Accept-Language": "ru",
        }

    async def _get_session(self) -> aiohttp.ClientSession:
        if self._session is None or self._session.closed:
            connector = aiohttp.TCPConnector(limit=100, limit_per_host=30, enable_cleanup_closed=True)
            timeout = aiohttp.ClientTimeout(total=20, connect=5)
            self._session = aiohttp.ClientSession(connector=connector, timeout=timeout, headers=self._headers)
        return self._session

    async def close(self):
        if self._session and not self._session.closed:
            await self._session.close()

    def _cleanup_balance_cache(self):
        now = datetime.now()
        expired = [k for k, v in self._balance_cache.items() if (now - v[1]).total_seconds() > 60]
        for k in expired:
            self._balance_cache.pop(k, None)

    async def get_all_drivers(self, force_refresh: bool = False) -> Tuple[List[dict], str]:
        if not self._is_configured():
            return [], "Yandex API kalitlari .env da to'liq emas!"

        now = datetime.now()
        if (not force_refresh and self._drivers_cache and self._cache_ts
                and (now - self._cache_ts).total_seconds() < self._cache_ttl):
            return self._drivers_cache, ""

        url = f"{self.FLEET_BASE}/v1/parks/driver-profiles/list"
        all_drivers: List[dict] = []
        limit, offset = 500, 0
        last_error = ""

        try:
            session = await self._get_session()
            while True:
                payload = {
                    "query": {
                        "park": {"id": self.park_id}
                    },
                    "limit": limit,
                    "offset": offset
                }
                async with session.post(url, json=payload) as resp:
                    text = await resp.text()
                    if resp.status != 200:
                        last_error = f"HTTP {resp.status}: {text[:250]}"
                        logger.error(f"Yandex get_all_drivers xatosi: {last_error}")
                        break

                    data = json.loads(text)
                    batch = data.get("driver_profiles", [])
                    all_drivers.extend(batch)
                    if len(batch) < limit:
                        break
                    offset += limit
        except Exception as e:
            last_error = f"Ulanish xatosi: {str(e)}"
            logger.error(f"Yandex get_all_drivers exception: {e}")

        if all_drivers:
            self._drivers_cache = all_drivers
            self._cache_ts = now
            logger.info(f"Yandex Fleet: {len(all_drivers)} ta haydovchi yangilandi.")
            return all_drivers, ""

        return [], last_error

    async def get_driver_by_phone(self, phone: str) -> Optional[dict]:
        if not self._is_configured():
            return None
        clean_target = clean_phone_number(phone)
        digits_target = re.sub(r"\D", "", clean_target)
        short9 = digits_target[-9:] if len(digits_target) >= 9 else digits_target

        drivers, _ = await self.get_all_drivers(force_refresh=True)
        for raw in drivers:
            prof = raw.get("driver_profile", {})
            for p in prof.get("phones", []):
                p_digits = re.sub(r"\D", "", str(p))
                if short9 and p_digits.endswith(short9):
                    return self._normalize(raw)
        return None

    def _extract_balance(self, raw_driver: dict) -> int:
        accounts = raw_driver.get("accounts", [])
        if not accounts:
            return 0
        for acc in accounts:
            if acc.get("type", "").lower() in ("personal_wallet", "wallet", "main"):
                try:
                    return int(float(acc.get("balance", 0)))
                except Exception:
                    pass
        try:
            return int(float(accounts[0].get("balance", 0)))
        except Exception:
            return 0

    def _normalize(self, raw_driver: dict) -> dict:
        prof = raw_driver.get("driver_profile", {})
        car = raw_driver.get("car", {})
        last = prof.get("last_name", "").strip()
        first = prof.get("first_name", "").strip()
        middle = prof.get("middle_name", "").strip()
        full_name = f"{last} {first} {middle}".strip() or "Haydovchi"
        
        brand_model = car.get("brand_and_model", "").strip()
        if not brand_model:
            brand_model = f"{car.get('brand','').strip()} {car.get('model','').strip()}".strip()
        car_title = brand_model or "Chevrolet Cobalt"
        car_number = car.get("number", "").strip() or car.get("normalized_number", "").strip() or "Noma'lum"
        phones = prof.get("phones", [])
        phone = clean_phone_number(phones[0]) if phones else ""

        return {
            "id": prof.get("id", ""),
            "full_name": full_name,
            "phone": phone,
            "car_model": car_title,
            "car_number": car_number,
            "balance": self._extract_balance(raw_driver),
            "raw": raw_driver,
        }

    async def get_driver_balance(self, yandex_driver_id: str) -> Optional[int]:
        if not self._is_configured() or not yandex_driver_id:
            return None

        self._cleanup_balance_cache()
        now = datetime.now()
        if yandex_driver_id in self._balance_cache:
            cached_bal, cached_time = self._balance_cache[yandex_driver_id]
            if (now - cached_time).total_seconds() < 10:
                return cached_bal

        url = f"{self.FLEET_BASE}/v1/parks/driver-profiles/list"
        payload = {
            "query": {
                "park": {"id": self.park_id},
                "driver": {"id": [yandex_driver_id]}
            },
            "limit": 1
        }
        try:
            session = await self._get_session()
            async with session.post(url, json=payload) as resp:
                text = await resp.text()
                if resp.status == 200:
                    data = json.loads(text)
                    drivers = data.get("driver_profiles", [])
                    if drivers:
                        bal = self._extract_balance(drivers[0])
                        self._balance_cache[yandex_driver_id] = (bal, now)
                        return bal
        except Exception as e:
            logger.error(f"Yandex get_driver_balance xatosi: {e}")
        return None

    async def get_today_orders_stats(self, yandex_driver_id: str) -> dict:
        default_res = {
            "total_orders": 0, "completed_orders": 0, "cancelled_orders": 0,
            "total_earnings": 0, "cash_earnings": 0, "card_earnings": 0, "park_comm": 0
        }
        if not self._is_configured() or not yandex_driver_id:
            return default_res

        now_tashkent = datetime.now(TASHKENT_TZ)
        today_start = now_tashkent.replace(hour=0, minute=0, second=0, microsecond=0)

        url = f"{self.FLEET_BASE}/v1/parks/orders/list"
        payload = {
            "query": {
                "park": {
                    "id": self.park_id,
                    "order": {
                        "booked_at": {
                            "from": today_start.isoformat(),
                            "to": now_tashkent.isoformat()
                        },
                        "driver_profile_id": yandex_driver_id
                    }
                }
            },
            "limit": 500
        }
        try:
            session = await self._get_session()
            async with session.post(url, json=payload) as resp:
                text = await resp.text()
                if resp.status == 200:
                    data = json.loads(text)
                    orders = data.get("orders", [])
                    comp = canc = 0
                    total_sum = cash_sum = card_sum = 0
                    for o in orders:
                        st = o.get("status", "").lower()
                        cost = int(float(o.get("cost", 0) or 0))
                        pay_type = o.get("payment_method", "card").lower()
                        if st == "complete":
                            comp += 1
                            total_sum += cost
                            if "cash" in pay_type:
                                cash_sum += cost
                            else:
                                card_sum += cost
                        elif st in ("cancelled", "canceled"):
                            canc += 1
                    comm = int(total_sum * (COMMISSION_PERCENT / 100.0))
                    return {
                        "total_orders": len(orders),
                        "completed_orders": comp,
                        "cancelled_orders": canc,
                        "total_earnings": total_sum,
                        "cash_earnings": cash_sum,
                        "card_earnings": card_sum,
                        "park_comm": comm
                    }
        except Exception as e:
            logger.error(f"Yandex get_today_orders_stats xatosi: {e}")
        return default_res

    async def create_transaction(self, yandex_driver_id: str, amount: int, description: str) -> bool:
        if not self._is_configured() or not yandex_driver_id:
            return False
        url = f"{self.FLEET_BASE}/v1/parks/driver-profiles/transactions"
        payload = {
            "park_id": self.park_id,
            "driver_profile_id": yandex_driver_id,
            "amount": str(-abs(int(amount))),
            "category_id": "other",
            "description": description,
        }
        try:
            session = await self._get_session()
            async with session.post(url, json=payload) as resp:
                text = await resp.text()
                if resp.status in (200, 201):
                    logger.info(f"Yandex tranzaksiya OK: Driver={yandex_driver_id}, Summa=-{amount}")
                    return True
                logger.error(f"Yandex tranzaksiya XATO: HTTP {resp.status} | {text[:200]}")
        except Exception as e:
            logger.error(f"Yandex tranzaksiya exception: {e}")
        return False


yandex_api = YandexFleetAPI(YANDEX_API_KEY, YANDEX_CLIENT_ID, YANDEX_PARK_ID)


# ============================================================
# 5. EXCEL HISOBOT
# ============================================================

async def generate_monthly_excel_report() -> bytes:
    drivers = await db_get_all_registered_drivers()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lochin Taxi Hisoboti"
    
    headers = [
        "№", "POSITION", "F.I.O (Haydovchi)", "Telefon Raqam", "Avtomobil Rusumi",
        "Davlat Raqami", "Plastik Karta (Maskalangan)", "Jami Buyurtmalar", 
        "Jami Daromad (so'm)", "Komissiya (so'm)", "Joriy Balans (so'm)", "Yandex Driver ID"
    ]
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
    )
    
    ws.append(headers)
    ws.row_dimensions[1].height = 28
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill, cell.font, cell.alignment = header_fill, header_font, align_center

    total_orders = total_earn = total_comm_sum = total_bal = 0
    
    for idx, drv in enumerate(drivers, 1):
        orders = int(drv.get("total_orders", 0) or 0)
        bal = int(drv.get("balance", 0) or 0)
        earnings = int(drv.get("total_earnings", 0) or 0)
        if earnings <= 0 and bal > 0:
            earnings = int(bal * 1.15)
        comm = int(earnings * (COMMISSION_PERCENT / 100.0))
        
        total_orders += orders
        total_earn += earnings
        total_comm_sum += comm
        total_bal += bal
        
        masked_card_val = mask_card(drv.get("card_number", ""))
        
        ws.append([
            idx,
            drv.get("position") or "N/A",
            drv.get("full_name") or "Noma'lum",
            drv.get("phone", ""),
            drv.get("car_model", ""),
            drv.get("car_number", ""),
            masked_card_val,
            orders,
            earnings,
            comm,
            bal,
            drv.get("yandex_driver_id") or "Yo'q"
        ])
        
        row_idx = idx + 1
        ws.row_dimensions[row_idx].height = 20
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = thin_border
            if col_num in [1, 2, 6, 7]:
                cell.alignment = align_center
            elif col_num in [8, 9, 10, 11]:
                cell.alignment = align_right
                cell.number_format = "#,##0"
            else:
                cell.alignment = align_left

    last_row = len(drivers) + 2
    ws.append([
        "JAMI", "", f"{len(drivers)} ta haydovchi", "", "", "", "",
        total_orders, total_earn, total_comm_sum, total_bal, ""
    ])
    ws.row_dimensions[last_row].height = 24
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=last_row, column=col_num)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.border = thin_border
        cell.alignment = align_right if col_num in [8, 9, 10, 11] else align_center
        if col_num in [8, 9, 10, 11]:
            cell.number_format = "#,##0"

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================
# 6. MATNLAR VA KLAVIATURALAR
# ============================================================

TEXTS = {
    "uz": {
        "welcome": f"🕌 <b>Assalomu alaykum!</b>\n\n🚕 <b>{BOT_NAME}</b> taksoparkiga xush kelibsiz! Biz bilan daromadingizni oshiring! 🤝\n\nTizimdan to'liq foydalanish uchun ro'yxatdan o'ting:",
        "register_btn": "📝 Ro'yxatdan o'tish",
        "reg_phone": "📱 <b>Telefon raqamingizni tasdiqlang:</b>\n\nXavfsizlik yuzasidan quyidagi <b>[📱 Telefon raqamni yuborish]</b> tugmasini bosing:",
        "reg_card": "💳 <b>Plastik karta raqamingizni kiriting (16 ta raqam):</b>\n\n<i>Misol: 8600 1234 5678 9012 yoki 9860...</i>",
        "reg_name": "👤 <b>Ism va familiyangizni kiriting:</b>\n\n<i>Misol: Alisher Qodirov</i>",
        "reg_car_model": "🚗 <b>Avtomobilingiz rusumini kiriting:</b>\n\n<i>Misol: Chevrolet Cobalt</i>",
        "reg_car_number": "🔢 <b>Avtomobil davlat raqamini kiriting:</b>\n\n<i>Misol: 01 A 123 AA</i>",
        "reg_success": "✅ <b>Tabriklaymiz! Siz muvaffaqiyatli ro'yxatdan o'tdingiz.</b>\n\n🆔 Sizning POSITION ID: <code>{position}</code>\n🔑 Bu kod sizning taksoparkdagi shaxsiy kodingiz.",
        "already_reg": "✅ <b>Siz tizimda ro'yxatdan o'tgansiz!</b>\n\n🆔 POSITION: <code>{position}</code>\n👤 Haydovchi: <b>{name}</b>",
        "menu_balance": "💰 Balans",
        "menu_orders": "📊 Bugungi buyurtmalar",
        "menu_withdraw": "💸 Pul yechish (24/7)",
        "menu_profile": "👤 Profil",
        "menu_top": "🏆 TOP Haydovchilar",
        "menu_group": "📢 Yangiliklar / Guruh",
        "menu_sos": "🆘 Yordam / SOS",
        "menu_admin": "🛠 Admin Panel",
        "cancel": "❌ Bekor qilish",
        "send_phone_btn": "📱 Telefon raqamni yuborish",
        "action_cancelled": "❌ Amaliyot bekor qilindi.",
        "withdraw_no_money": f"❌ Balansingizda yetarli mablag' yo'q!\nMinimal depozit qolishi shart: <b>{fmt_sum(MIN_DEPOSIT)} so'm</b>",
        "withdraw_min_err": f"❌ Minimal yechish summasi: {fmt_sum(MIN_WITHDRAWAL)} so'm",
        "withdraw_ask": f"💸 <b>Pul yechish arizasi:</b>\n\n🔹 Yechish mumkin: <b>{{avail}} so'm</b>\n🔹 Min. depozit qolishi shart: <b>{fmt_sum(MIN_DEPOSIT)} so'm</b>\n\nYechmoqchi bo'lgan summani kiriting (Masalan: <i>50000</i>):",
        "sos_title": f"🆘 <b>Tezkor Yordam va Aloqa Markazi</b>\n\n📞 <b>Menejer:</b> {SUPPORT_PHONE_DISPLAY}",
        "sos_btn_loc": "📍 Lokatsiya yuborish (DTP / Yo'lda qoldim)",
        "sos_btn_msg": "✍️ Menejerga xabar yozish",
        "sos_btn_chat": "💬 Menejer bilan shaxsiy chat",
        "sos_ask_loc": "📍 <b>Lokatsiyangizni yuboring (yoki matn ko'rinishida yozing):</b>",
        "sos_loc_btn": "📍 Hozirgi joylashuvimni yuborish",
        "sos_ask_msg": "✍️ <b>Muammo yoki savolingizni yozib qoldiring:</b>",
        "sos_sent": "🚨 <b>Xabaringiz Bosh Menejerga yetkazildi!</b>",
    },
    "ru": {
        "welcome": f"🕌 <b>Ассаламу алейкум!</b>\n\n🚕 Добро пожаловать в таксопарк <b>{BOT_NAME}</b>! Увеличьте свой доход с нами! 🤝\n\nПройдите регистрацию:",
        "register_btn": "📝 Регистрация",
        "reg_phone": "📱 <b>Подтвердите ваш номер телефона:</b>\n\nНажмите кнопку <b>[📱 Отправить номер]</b> ниже:",
        "reg_card": "💳 <b>Введите 16-значный номер карты:</b>\n\n<i>Пример: 8600 1234 5678 9012</i>",
        "reg_name": "👤 <b>Введите имя и фамилию:</b>\n\n<i>Пример: Алишер Кадыров</i>",
        "reg_car_model": "🚗 <b>Введите марку авто:</b>\n\n<i>Пример: Chevrolet Cobalt</i>",
        "reg_car_number": "🔢 <b>Введите госномер авто:</b>\n\n<i>Пример: 01 A 123 AA</i>",
        "reg_success": "✅ <b>Вы успешно зарегистрированы.</b>\n\n🆔 Ваш POSITION ID: <code>{position}</code>",
        "already_reg": "✅ <b>Вы уже зарегистрированы!</b>\n\n🆔 POSITION: <code>{position}</code>\n👤 Водитель: <b>{name}</b>",
        "menu_balance": "💰 Баланс",
        "menu_orders": "📊 Сегодняшние заказы",
        "menu_withdraw": "💸 Вывод средств (24/7)",
        "menu_profile": "👤 Профиль",
        "menu_top": "🏆 ТОП Водителей",
        "menu_group": "📢 Новости / Группа",
        "menu_sos": "🆘 Помощь / SOS",
        "menu_admin": "🛠 Админ Панель",
        "cancel": "❌ Отмена",
        "send_phone_btn": "📱 Отправить номер телефона",
        "action_cancelled": "❌ Действие отменено.",
        "withdraw_no_money": f"❌ Недостаточно средств!\nМин. депозит: <b>{fmt_sum(MIN_DEPOSIT)} сум</b>",
        "withdraw_min_err": f"❌ Мин. сумма вывода: {fmt_sum(MIN_WITHDRAWAL)} сум",
        "withdraw_ask": f"💸 <b>Заявка на вывод:</b>\n\n🔹 Доступно: <b>{{avail}} сум</b>\n🔹 Мин. депозит: <b>{fmt_sum(MIN_DEPOSIT)} сум</b>\n\nВведите сумму для вывода:",
        "sos_title": f"🆘 <b>Центр Помощи</b>\n\n📞 <b>Телефон:</b> {SUPPORT_PHONE_DISPLAY}",
        "sos_btn_loc": "📍 Отправить локацию (ДТП)",
        "sos_btn_msg": "✍️ Написать менеджеру",
        "sos_btn_chat": "💬 Чат с менеджером",
        "sos_ask_loc": "📍 <b>Отправьте вашу локацию (или напишите текст):</b>",
        "sos_loc_btn": "📍 Отправить локацию",
        "sos_ask_msg": "✍️ <b>Опишите вопрос:</b>",
        "sos_sent": "🚨 <b>Сообщение отправлено!</b>",
    }
}


def t(lang_code: str, key: str, **kwargs) -> str:
    text = TEXTS.get(lang_code, TEXTS["uz"]).get(key, key)
    return text.format(**kwargs) if kwargs else text


def user_main_kb(lang: str, uid: int) -> ReplyKeyboardMarkup:
    buttons = [
        [KeyboardButton(text=t(lang, "menu_balance")), KeyboardButton(text=t(lang, "menu_withdraw"))],
        [KeyboardButton(text=t(lang, "menu_orders")), KeyboardButton(text=t(lang, "menu_profile"))],
        [KeyboardButton(text=t(lang, "menu_top")), KeyboardButton(text=t(lang, "menu_group"))],
        [KeyboardButton(text=t(lang, "menu_sos"))],
    ]
    if uid in ADMIN_IDS:
        buttons.append([KeyboardButton(text=t(lang, "menu_admin"))])
    return ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)


def cancel_kb(lang: str) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text=t(lang, "cancel"))]], resize_keyboard=True)


def phone_request_kb(lang: str) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=t(lang, "send_phone_btn"), request_contact=True)],
            [KeyboardButton(text=t(lang, "cancel"))]
        ],
        resize_keyboard=True,
    )


def location_request_kb(lang: str) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=t(lang, "sos_loc_btn"), request_location=True)],
            [KeyboardButton(text=t(lang, "cancel"))]
        ],
        resize_keyboard=True,
    )


def language_inline_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="🇺🇿 O'zbekcha", callback_data="lang:uz"),
        InlineKeyboardButton(text="🇷🇺 Русский", callback_data="lang:ru"),
    ]])


def register_reply_kb(lang: str) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text=t(lang, "register_btn"))]], resize_keyboard=True)


def sos_menu_kb(lang: str) -> InlineKeyboardMarkup:
    kb_rows = [
        [InlineKeyboardButton(text=t(lang, "sos_btn_loc"), callback_data="sos:loc")],
        [InlineKeyboardButton(text=t(lang, "sos_btn_msg"), callback_data="sos:msg")],
    ]
    if MANAGER_TG_ID > 0:
        kb_rows.append([InlineKeyboardButton(text=t(lang, "sos_btn_chat"), url=f"tg://user?id={MANAGER_TG_ID}")])
    return InlineKeyboardMarkup(inline_keyboard=kb_rows)


def admin_main_kb(lang: str) -> ReplyKeyboardMarkup:
    is_uz = lang == "uz"
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="📊 Statistika" if is_uz else "📊 Статистика"),
         KeyboardButton(text="📥 Excel Hisobot" if is_uz else "📥 Excel Отчет")],
        [KeyboardButton(text="🔄 Yandex Sinxronlash" if is_uz else "🔄 Синхронизация Яндекс"),
         KeyboardButton(text="📢 Xabar tarqatish" if is_uz else "📢 Рассылка")],
        [KeyboardButton(text="👥 Haydovchilar" if is_uz else "👥 Водители"),
         KeyboardButton(text="🗑 Haydovchini o'chirish" if is_uz else "🗑 Удалить водителя")],
        [KeyboardButton(text="🚫 Nofaollar" if is_uz else "🚫 Неактивные"),
         KeyboardButton(text="⬅️ Asosiy menyu" if is_uz else "⬅️ Главное меню")],
    ], resize_keyboard=True)


# ============================================================
# 7. ANTI-FLOOD THROTTLING
# ============================================================

class ThrottlingMiddleware(BaseMiddleware):
    def __init__(self, limit: float = 0.4):
        self.limit = limit
        self.user_timestamps: Dict[int, float] = {}
        self.last_cleanup = time.time()

    def _cleanup_old_entries(self, now: float):
        if now - self.last_cleanup > 3600 or len(self.user_timestamps) > 2000:
            threshold = now - 60.0
            self.user_timestamps = {uid: ts for uid, ts in self.user_timestamps.items() if ts > threshold}
            self.last_cleanup = now

    async def __call__(
        self,
        handler: Callable[[TelegramObject, Dict[str, Any]], Awaitable[Any]],
        event: TelegramObject,
        data: Dict[str, Any]
    ) -> Any:
        if isinstance(event, Message) and event.from_user:
            user_id = event.from_user.id
            now = time.time()
            self._cleanup_old_entries(now)
            
            last_time = self.user_timestamps.get(user_id, 0.0)
            if now - last_time < self.limit:
                return
            self.user_timestamps[user_id] = now
        return await handler(event, data)


# ============================================================
# 8. FSM STATES
# ============================================================

class RegStates(StatesGroup):
    phone = State()
    name = State()
    card = State()
    car_model = State()
    car_number = State()


class WithdrawStates(StatesGroup):
    amount = State()
    confirm = State()


class SOSStates(StatesGroup):
    waiting_for_location = State()
    waiting_for_message = State()


class AdminBroadcastStates(StatesGroup):
    waiting_for_message = State()


class AdminDeleteDriverStates(StatesGroup):
    waiting_for_query = State()


# ============================================================
# 9. DISPATCHER VA ROUTERLAR
# ============================================================

bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())
dp.message.middleware(ThrottlingMiddleware(limit=0.4))

router = Router()
admin_router = Router()

CANCEL_TEXTS = {"❌ Bekor qilish", "❌ Отмена", "bekor", "отмена", "/bekor", "/cancel"}


async def get_lang(uid: int) -> str:
    user = await db_get_user(uid)
    return user.get("language", "uz") if user else "uz"


@router.message(Command("cancel"), StateFilter("*"))
@router.message(F.text.in_(CANCEL_TEXTS), StateFilter("*"))
async def global_cancel_handler(message: Message, state: FSMContext) -> None:
    await state.clear()
    uid = message.from_user.id
    lang = await get_lang(uid)
    user = await db_get_user(uid)
    kb = user_main_kb(lang, uid) if (user and user.get("is_registered") == 1) else register_reply_kb(lang)
    await message.answer(t(lang, "action_cancelled"), reply_markup=kb)


@router.message(CommandStart(), StateFilter("*"))
async def cmd_start(message: Message, state: FSMContext) -> None:
    await state.clear()
    uid = message.from_user.id
    await db_upsert_start(uid, message.from_user.username or "")
    user = await db_get_user(uid)
    if user and user.get("is_registered") == 1:
        lang = user.get("language", "uz")
        pos_id = user.get("position") or "N/A"
        drv_name = user.get("full_name") or "Haydovchi"
        await message.answer(
            t(lang, "already_reg", position=pos_id, name=drv_name),
            reply_markup=user_main_kb(lang, uid),
        )
        return
    await message.answer("🌐 <b>Iltimos, tilni tanlang / Пожалуйста, выберите язык:</b>", reply_markup=language_inline_kb())


@router.callback_query(F.data.startswith("lang:"))
async def lang_callback(callback: CallbackQuery) -> None:
    lang = callback.data.split(":")[1]
    uid = callback.from_user.id
    await db_set_language(uid, lang)
    try:
        await callback.message.delete()
    except Exception:
        pass
    user = await db_get_user(uid)
    if user and user.get("is_registered") == 1:
        pos_id = user.get("position") or "N/A"
        drv_name = user.get("full_name") or "Haydovchi"
        await callback.message.answer(
            t(lang, "already_reg", position=pos_id, name=drv_name),
            reply_markup=user_main_kb(lang, uid),
        )
    else:
        await callback.message.answer(t(lang, "welcome"), reply_markup=register_reply_kb(lang))
    await callback.answer()


# ============================================================
# 10. RO'YXATDAN O'TISH HANDLERLARI
# ============================================================

@router.message(F.text.in_(["📝 Ro'yxatdan o'tish", "📝 Регистрация"]), StateFilter("*"))
async def reg_start_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    uid = message.from_user.id
    user = await db_get_user(uid)
    lang = user.get("language", "uz") if user else "uz"
    if user and user.get("is_registered") == 1:
        pos_id = user.get("position") or "N/A"
        drv_name = user.get("full_name") or "Haydovchi"
        await message.answer(
            t(lang, "already_reg", position=pos_id, name=drv_name),
            reply_markup=user_main_kb(lang, uid),
        )
        return
    await state.set_state(RegStates.phone)
    await message.answer(t(lang, "reg_phone"), reply_markup=phone_request_kb(lang))


@router.message(RegStates.phone)
async def reg_step_phone(message: Message, state: FSMContext) -> None:
    uid = message.from_user.id
    lang = await get_lang(uid)

    if not message.contact:
        await message.answer(
            "⚠️ <b>Xavfsizlik talabi:</b> Iltimos, pastdagi <b>[📱 Telefon raqamni yuborish]</b> tugmasini bosing.\nQo'lda yozilgan raqamlar qabul qilinmaydi!",
            reply_markup=phone_request_kb(lang)
        )
        return

    if message.contact.user_id != uid:
        await message.answer(
            "❌ <b>Xatolik!</b> Siz faqat o'zingizning Telegram profilingizga tegishli telefon raqamni yuborishingiz mumkin.",
            reply_markup=phone_request_kb(lang)
        )
        return

    phone = clean_phone_number(message.contact.phone_number)

    existing_phone_user = await db_get_user_by_phone(phone)
    if existing_phone_user and existing_phone_user.get("telegram_id") != uid and existing_phone_user.get("is_registered") == 1:
        await message.answer(
            "❌ <b>Ushbu telefon raqami allaqachon boshqa profilga biriktirilgan!</b>\nIltimos, ma'muriyatga murojaat qiling.",
            reply_markup=user_main_kb(lang, uid)
        )
        await state.clear()
        return

    await state.update_data(phone=phone)
    search_msg = await message.answer("⏳ <i>Yandex bazasidan haydovchi tekshirilmoqda...</i>")
    
    y_driver = await yandex_api.get_driver_by_phone(phone)
    try:
        await search_msg.delete()
    except Exception:
        pass

    if y_driver:
        await state.update_data(
            full_name=y_driver.get("full_name") or "Haydovchi",
            car_model=y_driver.get("car_model") or "Chevrolet Cobalt",
            car_number=y_driver.get("car_number") or "Noma'lum",
            yandex_driver_id=y_driver.get("id"),
        )
        drv_nm = y_driver.get("full_name", "")
        car_md = y_driver.get("car_model", "")
        car_nb = y_driver.get("car_number", "")
        card_prompt = t(lang, "reg_card")
        found_txt = (
            f"✅ <b>Siz Yandex Pro taksoparkimizda topildingiz!</b>\n\n"
            f"👤 Haydovchi: <b>{drv_nm}</b>\n"
            f"🚗 Avtomobil: <b>{car_md} ({car_nb})</b>\n\n"
            f"{card_prompt}"
        )
        await state.set_state(RegStates.card)
        await message.answer(found_txt, reply_markup=cancel_kb(lang))
    else:
        await state.set_state(RegStates.name)
        await message.answer(t(lang, "reg_name"), reply_markup=cancel_kb(lang))


@router.message(RegStates.name)
async def reg_step_name(message: Message, state: FSMContext) -> None:
    uid = message.from_user.id
    lang = await get_lang(uid)
    name = (message.text or "").strip()
    if len(name) < 3:
        await message.answer("⚠️ Iltimos, ism va familiyangizni to'liq kiriting:")
        return
    await state.update_data(full_name=name)
    await state.set_state(RegStates.card)
    await message.answer(t(lang, "reg_card"), reply_markup=cancel_kb(lang))


@router.message(RegStates.card)
async def reg_step_card(message: Message, state: FSMContext) -> None:
    uid = message.from_user.id
    lang = await get_lang(uid)
    card = re.sub(r"\D", "", message.text or "")
    if not (card.isdigit() and len(card) == 16):
        await message.answer("⚠️ Plastik karta aynan 16 ta raqamdan iborat bo'lishi kerak:")
        return
    await state.update_data(card_number=card)
    data = await state.get_data()
    if data.get("yandex_driver_id"):
        await finish_registration_process(message, state, data)
    else:
        await state.set_state(RegStates.car_model)
        await message.answer(t(lang, "reg_car_model"), reply_markup=cancel_kb(lang))


@router.message(RegStates.car_model)
async def reg_step_car_model(message: Message, state: FSMContext) -> None:
    await state.update_data(car_model=(message.text or "").strip())
    lang = await get_lang(message.from_user.id)
    await state.set_state(RegStates.car_number)
    await message.answer(t(lang, "reg_car_number"), reply_markup=cancel_kb(lang))


@router.message(RegStates.car_number)
async def reg_step_car_number(message: Message, state: FSMContext) -> None:
    await state.update_data(car_number=(message.text or "").strip().upper())
    data = await state.get_data()
    await finish_registration_process(message, state, data)


async def finish_registration_process(message: Message, state: FSMContext, data: dict):
    uid = message.from_user.id
    lang = await get_lang(uid)
    await state.clear()
    full_name = data.get("full_name") or "Haydovchi"
    phone = data.get("phone") or ""
    card = data.get("card_number") or ""
    car_model = data.get("car_model") or "Chevrolet"
    car_number = data.get("car_number") or ""
    y_id = data.get("yandex_driver_id")

    position = await db_finish_registration(
        telegram_id=uid, full_name=full_name, phone=phone, card_number=card,
        car_model=car_model, car_number=car_number, yandex_driver_id=y_id,
    )

    await message.answer(t(lang, "reg_success", position=position), reply_markup=user_main_kb(lang, uid))

    init_bal = 0
    if y_id:
        live_b = await yandex_api.get_driver_balance(y_id)
        if live_b is not None:
            init_bal = live_b
            await db_update_balance(uid, live_b)

    yandex_txt = "Ulangan ✅" if y_id else "Ulanmagan ❌"
    
    admin_alert = (
        f"🆕 <b>YANGI HAYDOVCHI RO'YXATDAN O'TDI!</b>\n\n"
        f"🆔 POSITION: <code>{position}</code>\n"
        f"👤 <b>Haydovchi:</b> {full_name}\n"
        f"📱 <b>Telefon:</b> <code>{phone}</code>\n"
        f"🚗 <b>Avtomobil:</b> {car_model} ({car_number})\n"
        f"💳 <b>Karta (Maskalangan):</b> <code>{mask_card(card)}</code>\n"
        f"💰 <b>Boshlang'ich Balans:</b> <b>{fmt_sum(init_bal)} so'm</b>\n"
        f"🚖 <b>Yandex Pro:</b> {yandex_txt}"
    )
    
    adm_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="💬 Haydovchi bilan chat", url=f"tg://user?id={uid}")
    ]])

    for adm in ADMIN_IDS:
        try:
            await bot.send_message(adm, admin_alert, reply_markup=adm_kb)
        except Exception:
            pass


# ============================================================
# 11. BALANS (10s KESH & MASKALANGAN KARTA)
# ============================================================

@router.message(F.text.in_(["💰 Balans", "💰 Баланс"]))
async def balance_handler(message: Message) -> None:
    uid = message.from_user.id
    user = await db_get_user(uid)
    if not user or user.get("is_registered") != 1:
        await message.answer("Iltimos, avval ro'yxatdan o'ting: /start", reply_markup=register_reply_kb("uz"))
        return

    lang = user.get("language", "uz")
    cur_bal = int(user.get("balance", 0) or 0)
    y_status = "Ulangan ✅" if user.get("yandex_driver_id") else "Ulanmagan ❌"

    if user.get("yandex_driver_id"):
        live_bal = await yandex_api.get_driver_balance(user["yandex_driver_id"])
        if live_bal is not None:
            cur_bal = live_bal
            await db_update_balance(uid, live_bal)

    today_withdrawn = await db_get_driver_today_withdrawn(user["id"])
    avail = max(0, cur_bal - MIN_DEPOSIT)

    u_pos = user.get("position", "N/A")
    u_name = user.get("full_name", "Haydovchi")
    u_phone = user.get("phone", "")
    u_car = f"{user.get('car_model','')} ({user.get('car_number','')})"
    u_card_masked = mask_card(user.get("card_number", ""))

    text = (
        f"💰 <b>{BOT_NAME} — Shaxsiy Balans va Ma'lumot:</b>\n\n"
        f"👤 <b>Haydovchi:</b> {u_name}\n"
        f"🆔 <b>POSITION:</b> <code>{u_pos}</code>\n"
        f"📱 <b>Telefon:</b> <code>{u_phone}</code>\n"
        f"🚗 <b>Avtomobil:</b> {u_car}\n"
        f"💳 <b>Karta:</b> <code>{u_card_masked}</code>\n"
        f"➖➖➖➖➖➖➖➖➖➖\n"
        f"💳 <b>Yandex Pro Balans:</b> <b>{fmt_sum(cur_bal)} so'm</b>\n"
        f"🔒 <b>Minimal depozit (majburiy):</b> {fmt_sum(MIN_DEPOSIT)} so'm\n"
        f"💸 <b>Bugun yechib olingan:</b> <b>{fmt_sum(today_withdrawn)} so'm</b>\n"
        f"➖➖➖➖➖➖➖➖➖➖\n"
        f"✅ <b>Kartaga yechish mumkin:</b> <b>{fmt_sum(avail)} so'm</b>\n\n"
        f"🚕 <b>Yandex Pro holati:</b> {y_status}"
    ) if lang == "uz" else (
        f"💰 <b>{BOT_NAME} — Личный Баланс и Данные:</b>\n\n"
        f"👤 <b>Водитель:</b> {u_name}\n"
        f"🆔 <b>POSITION:</b> <code>{u_pos}</code>\n"
        f"📱 <b>Телефон:</b> <code>{u_phone}</code>\n"
        f"🚗 <b>Автомобиль:</b> {u_car}\n"
        f"💳 <b>Карта:</b> <code>{u_card_masked}</code>\n"
        f"➖➖➖➖➖➖➖➖➖➖\n"
        f"💳 <b>Баланс Яндекс Про:</b> <b>{fmt_sum(cur_bal)} сум</b>\n"
        f"🔒 <b>Неснижаемый остаток:</b> {fmt_sum(MIN_DEPOSIT)} сум\n"
        f"💸 <b>Выведено за сегодня:</b> <b>{fmt_sum(today_withdrawn)} сум</b>\n"
        f"➖➖➖➖➖➖➖➖➖➖\n"
        f"✅ <b>Доступно к выводу:</b> <b>{fmt_sum(avail)} сум</b>\n\n"
        f"🚕 <b>Статус Яндекс Про:</b> {y_status}"
    )
    await message.answer(text, reply_markup=user_main_kb(lang, uid))


# ============================================================
# 12. BUGUNGI BUYURTMALAR (REAL-TIME YANDEX)
# ============================================================

@router.message(F.text.in_(["📊 Bugungi buyurtmalar", "📊 Сегодняшние заказы"]))
async def orders_handler(message: Message) -> None:
    uid = message.from_user.id
    user = await db_get_user(uid)
    if not user or user.get("is_registered") != 1:
        await message.answer("Iltimos, avval ro'yxatdan o'ting: /start", reply_markup=register_reply_kb("uz"))
        return

    lang = user.get("language", "uz")
    wait_msg = await message.answer("⏳ <i>Yandex Pro dan bugungi buyurtmalar olinmoqda...</i>")

    y_id = user.get("yandex_driver_id")
    stats = await yandex_api.get_today_orders_stats(y_id) if y_id else {
        "total_orders": 0, "completed_orders": 0, "cancelled_orders": 0,
        "total_earnings": 0, "cash_earnings": 0, "card_earnings": 0, "park_comm": 0
    }
    try:
        await wait_msg.delete()
    except Exception:
        pass

    now_tashkent = datetime.now(TASHKENT_TZ)
    t_orders = stats.get("total_orders", 0)
    c_orders = stats.get("completed_orders", 0)
    x_orders = stats.get("cancelled_orders", 0)
    t_earn = fmt_sum(stats.get("total_earnings", 0))
    cd_earn = fmt_sum(stats.get("card_earnings", 0))
    cs_earn = fmt_sum(stats.get("cash_earnings", 0))
    p_comm = fmt_sum(stats.get("park_comm", 0))

    text = (
        f"📊 <b>Bugungi Buyurtmalar Statistikasi:</b>\n"
        f"📅 <i>{now_tashkent.strftime('%d.%m.%Y | %H:%M')} holatiga</i>\n\n"
        f"🚕 <b>Jami buyurtmalar:</b> <b>{t_orders} ta</b>\n"
        f"  └ ✅ Tugallangan: <b>{c_orders} ta</b>\n"
        f"  └ ❌ Bekor qilingan: <b>{x_orders} ta</b>\n\n"
        f"💰 <b>Bugungi jami daromad:</b> <b>{t_earn} so'm</b>\n"
        f"  └ 💳 Karta orqali: <b>{cd_earn} so'm</b>\n"
        f"  └ 💵 Naqd orqali: <b>{cs_earn} so'm</b>\n"
        f"📈 <b>Taksopark komissiyasi:</b> {p_comm} so'm\n"
        f"➖➖➖➖➖➖➖➖➖➖\n"
        f"🔥 <i>Buyurtmalarni faol bajaring va haftalik TOP mukofotlarga ega bo'ling!</i>"
    )
    await message.answer(text, reply_markup=user_main_kb(lang, uid))


# ============================================================
# 13. PUL YECHISH (1 TA PENDING ARIZA CHEKLOVI)
# ============================================================

@router.message(F.text.in_(["💸 Pul yechish (24/7)", "💸 Вывод средств (24/7)"]), StateFilter("*"))
async def withdraw_start(message: Message, state: FSMContext) -> None:
    await state.clear()
    uid = message.from_user.id
    user = await db_get_user(uid)
    if not user or user.get("is_registered") != 1:
        await message.answer("Iltimos, avval ro'yxatdan o'ting: /start", reply_markup=register_reply_kb("uz"))
        return

    lang = user.get("language", "uz")
    if user.get("yandex_driver_id"):
        live_bal = await yandex_api.get_driver_balance(user["yandex_driver_id"])
        if live_bal is not None:
            await db_update_balance(uid, live_bal)
            user["balance"] = live_bal

    cur_bal = int(user.get("balance", 0) or 0)
    avail = max(0, cur_bal - MIN_DEPOSIT)

    if avail < MIN_WITHDRAWAL:
        await message.answer(
            f"{t(lang, 'withdraw_no_money')}\n\n"
            f"🔹 Joriy balans: <b>{fmt_sum(cur_bal)} so'm</b>\n"
            f"🔹 Yechish mumkin: <b>{fmt_sum(avail)} so'm</b>\n"
            f"🔹 Minimal yechish: <b>{fmt_sum(MIN_WITHDRAWAL)} so'm</b>",
            reply_markup=user_main_kb(lang, uid)
        )
        return

    await state.set_state(WithdrawStates.amount)
    await message.answer(t(lang, "withdraw_ask", avail=fmt_sum(avail)), reply_markup=cancel_kb(lang))


@router.message(WithdrawStates.amount)
async def withdraw_amount_step(message: Message, state: FSMContext) -> None:
    uid = message.from_user.id
    lang = await get_lang(uid)
    raw = (message.text or "").replace(" ", "").replace("so'm", "").replace("som", "").replace("сум", "").strip()
    if not raw.isdigit():
        await message.answer("⚠️ Iltimos, summani faqat musbat butun raqamlarda kiriting:")
        return

    amount = int(raw)
    user = await db_get_user(uid)
    cur_bal = int(user.get("balance", 0) or 0)
    avail = max(0, cur_bal - MIN_DEPOSIT)

    if amount < MIN_WITHDRAWAL:
        await message.answer(t(lang, "withdraw_min_err"))
        return
    if amount > avail:
        await message.answer(f"❌ Mablag' yetarli emas! Siz ko'pi bilan <b>{fmt_sum(avail)} so'm</b> yecha olasiz.")
        return

    comm = int(amount * (COMMISSION_PERCENT / 100.0))
    net = amount - comm
    full_card_val = user.get("card_number") or ""
    masked_card_val = mask_card(full_card_val)
    rem_deposit = cur_bal - amount

    await state.update_data(amount=amount, commission=comm, net_amount=net, card=full_card_val, rem_deposit=rem_deposit)
    await state.set_state(WithdrawStates.confirm)

    confirm_txt = (
        f"💳 <b>Pul yechishni tasdiqlaysizmi?</b>\n\n"
        f"💰 Yechilayotgan summa: <b>{fmt_sum(amount)} so'm</b>\n"
        f"💵 Kartaga to'lanadi: <b>{fmt_sum(net)} so'm</b>\n"
        f"🔒 Depozitda qoladi: <b>{fmt_sum(rem_deposit)} so'm</b>\n"
        f"💳 Karta: <code>{masked_card_val}</code>"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Tasdiqlayman", callback_data="wd_go:yes"),
        InlineKeyboardButton(text="❌ Bekor qilish", callback_data="wd_go:no"),
    ]])
    await message.answer(confirm_txt, reply_markup=kb)


@router.callback_query(F.data.startswith("wd_go:"), WithdrawStates.confirm)
async def withdraw_process_callback(callback: CallbackQuery, state: FSMContext) -> None:
    uid = callback.from_user.id
    action = callback.data.split(":")[1]

    if action == "no":
        await state.clear()
        await callback.message.edit_text("❌ Pul yechish bekor qilindi.")
        await callback.answer()
        return

    data = await state.get_data()
    amount = data["amount"]
    commission = data["commission"]
    net_amount = data["net_amount"]
    full_card = data["card"]
    rem_deposit = data.get("rem_deposit", 0)
    await state.clear()

    user = await db_get_user(uid)
    if not user:
        await callback.message.edit_text("❌ Foydalanuvchi topilmadi.")
        await callback.answer()
        return

    try:
        w_id = await db_create_withdrawal(
            user_id=user["id"], telegram_id=uid, amount=amount, commission=commission,
            net_amount=net_amount, card_number=full_card, status="pending",
            payout_method="manual", ext_tx_id="",
        )
    except ValueError as val_err:
        await callback.message.edit_text(f"❌ Xatolik: {val_err}")
        await callback.answer("Amaliyot rad etildi!", show_alert=True)
        return

    masked_c = mask_card(full_card)
    await callback.message.edit_text(
        f"✅ <b>Pul yechish arizangiz qabul qilindi! (Ariza #{w_id})</b>\n\n"
        f"💰 Yechilayotgan summa: <b>{fmt_sum(amount)} so'm</b>\n"
        f"💵 Kartaga tushadi: <b>{fmt_sum(net_amount)} so'm</b>\n"
        f"💳 Karta: <code>{masked_c}</code>\n\n"
        f"⏱ <i>Mablag' qisqa vaqt ichida kartangizga o'tkaziladi.</i>"
    )
    await callback.answer()

    y_status_txt = "Ulangan ✅" if user.get("yandex_driver_id") else "Ulanmagan ❌"
    u_pos = user.get("position", "N/A")
    u_name = user.get("full_name", "")
    u_phone = user.get("phone", "")
    u_model = user.get("car_model", "")
    u_num = user.get("car_number", "")

    admin_alert = (
        f"💸 <b>YANGI PUL YECHISH ARIZASI! (Ariza #{w_id})</b>\n\n"
        f"🆔 POSITION: <code>{u_pos}</code>\n"
        f"👤 <b>Haydovchi:</b> {u_name}\n"
        f"📱 <b>Telefon:</b> <code>{u_phone}</code>\n"
        f"🚗 <b>Avtomobil:</b> {u_model} ({u_num})\n"
        f"💳 <b>Karta:</b> <code>{full_card}</code> <i>(Nusxa olish uchun bosing)</i>\n"
        f"➖➖➖➖➖➖➖➖➖➖\n"
        f"💰 <b>Yechilayotgan summa:</b> {fmt_sum(amount)} so'm\n"
        f"💵 <b>Kartaga to'lanadigan sof summa:</b> <b>{fmt_sum(net_amount)} so'm</b>\n"
        f"🔒 <b>Depozitda qoladigan:</b> {fmt_sum(rem_deposit)} so'm (Min. depozit: {fmt_sum(MIN_DEPOSIT)} so'm)\n"
        f"🚖 <b>Yandex Pro:</b> {y_status_txt}"
    )

    adm_kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ To'landi (Yandexdan yechish)", callback_data=f"adm_pay:{w_id}"),
            InlineKeyboardButton(text="❌ Rad etish", callback_data=f"adm_rej:{w_id}")
        ],
        [
            InlineKeyboardButton(text="💬 Haydovchi bilan chat", url=f"tg://user?id={uid}")
        ]
    ])

    for adm in ADMIN_IDS:
        try:
            log_admin_view_card(adm, w_id)
            await bot.send_message(adm, admin_alert, reply_markup=adm_kb)
        except Exception:
            pass


# ============================================================
# 14. ADMIN TASDIQLASH VA RAD ETISH
# ============================================================

@admin_router.callback_query(F.data.startswith("adm_pay:"))
async def admin_approve_payout(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Ruxsat yo'q!", show_alert=True)
        return

    w_id = int(callback.data.split(":")[1])
    wd = await db_get_withdrawal(w_id)
    if not wd or wd.get("status") != "pending":
        await callback.answer("Bu ariza allaqachon ko'rib chiqilgan!", show_alert=True)
        return

    user = await db_get_user_by_id(wd["user_id"])
    if not user:
        await callback.answer("Haydovchi topilmadi!", show_alert=True)
        return

    if user.get("yandex_driver_id"):
        await yandex_api.create_transaction(
            user["yandex_driver_id"],
            int(wd["amount"]),
            f"Lochin Taxi Bot to'lovi #{w_id} ({mask_card(wd.get('card_number',''))})"
        )

    await db_update_withdrawal_status(w_id, "completed")

    try:
        await callback.message.edit_text(
            f"{callback.message.text}\n\n✅ <b>TO'LANDI VA YANDEX PRODAN YECHILDI!</b>\n👨‍💻 Admin: {callback.from_user.full_name}"
        )
    except Exception:
        pass
    await callback.answer("To'lov tasdiqlandi!")

    try:
        await bot.send_message(
            user["telegram_id"],
            f"✅ <b>Pul yechish arizangiz tasdiqlandi! (Ariza #{w_id})</b>\n\n"
            f"💵 <b>{fmt_sum(wd['net_amount'])} so'm</b> kartangizga muvaffaqiyatli o'tkazildi.\n"
            f"💳 Karta: <code>{mask_card(wd.get('card_number',''))}</code>\n\n"
            f"<i>Lochin Taxi bilan ishlaganingiz uchun rahmat!</i>"
        )
    except Exception:
        pass


@admin_router.callback_query(F.data.startswith("adm_rej:"))
async def admin_reject_payout(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Ruxsat yo'q!", show_alert=True)
        return

    w_id = int(callback.data.split(":")[1])
    wd = await db_get_withdrawal(w_id)
    if not wd or wd.get("status") != "pending":
        await callback.answer("Bu ariza allaqachon ko'rib chiqilgan!", show_alert=True)
        return

    await db_refund_withdrawal(w_id)

    user = await db_get_user_by_id(wd["user_id"])
    try:
        await callback.message.edit_text(
            f"{callback.message.text}\n\n❌ <b>ARIZA RAD ETILDI VA BALANS QAYTARILDI.</b>\n👨‍💻 Admin: {callback.from_user.full_name}"
        )
    except Exception:
        pass
    await callback.answer("Ariza rad etildi!")

    if user:
        try:
            await bot.send_message(
                user["telegram_id"],
                f"❌ <b>Pul yechish arizangiz rad etildi. (Ariza #{w_id})</b>\n\n"
                f"💰 {fmt_sum(wd['amount'])} so'm balansingizga qaytarildi.\n"
                f"Batafsil ma'lumot uchun menejer bilan bog'laning."
            )
        except Exception:
            pass


# ============================================================
# 15. PROFIL, TOP, SOS
# ============================================================

@router.message(F.text.in_(["👤 Profil", "👤 Профиль"]))
async def profile_handler(message: Message) -> None:
    uid = message.from_user.id
    user = await db_get_user(uid)
    if not user or user.get("is_registered") != 1:
        await message.answer("Iltimos, avval ro'yxatdan o'ting: /start", reply_markup=register_reply_kb("uz"))
        return
    lang = user.get("language", "uz")
    y_val = "Ulangan ✅" if user.get("yandex_driver_id") else "Ulanmagan ❌"
    u_pos = user.get("position", "N/A")
    u_name = user.get("full_name", "")
    u_phone = user.get("phone", "")
    u_model = user.get("car_model", "")
    u_num = user.get("car_number", "")
    u_card_masked = mask_card(user.get("card_number", ""))
    lang_display = "O'zbekcha" if lang == "uz" else "Русский"

    text = (
        f"👤 <b>Haydovchi Profili:</b>\n\n"
        f"🆔 POSITION: <code>{u_pos}</code>\n"
        f"👤 Ism: <b>{u_name}</b>\n"
        f"📱 Telefon: <b>{u_phone}</b>\n"
        f"🚗 Avtomobil: <b>{u_model} ({u_num})</b>\n"
        f"💳 Karta: <code>{u_card_masked}</code>\n"
        f"🚕 Yandex: <b>{y_val}</b>\n"
        f"🌐 Til: <b>{lang_display}</b>"
    )
    btn_txt = "🌐 Tilni o'zgartirish" if lang == "uz" else "🌐 Сменить язык"
    await message.answer(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text=btn_txt, callback_data="change_lang_menu")]]))


@router.callback_query(F.data == "change_lang_menu")
async def change_lang_menu_cb(callback: CallbackQuery) -> None:
    await callback.message.edit_text("🌐 Tilni tanlang / Выберите язык:", reply_markup=language_inline_kb())
    await callback.answer()


@router.message(F.text.in_(["🏆 TOP Haydovchilar", "🏆 ТОП Водителей"]))
async def top_drivers_handler(message: Message) -> None:
    uid = message.from_user.id
    lang = await get_lang(uid)
    drivers = await db_get_all_registered_drivers()
    top = sorted(drivers, key=lambda x: int(x.get("total_orders", 0) or 0), reverse=True)[:10]
    medals = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣", "6️⃣", "7️⃣", "8️⃣", "9️⃣", "🔟"]
    text = f"🏆 <b>{BOT_NAME} — Haftaning Eng Faol Haydovchilari:</b>\n\n"
    if top and any(int(d.get("total_orders", 0) or 0) > 0 for d in top):
        for i, d in enumerate(top):
            medal_icon = medals[i] if i < len(medals) else str(i + 1)
            d_name = d.get("full_name", "")
            d_pos = d.get("position", "N/A")
            d_orders = d.get("total_orders", 0)
            text += f"{medal_icon}. <b>{d_name}</b> (<code>{d_pos}</code>) — <b>{d_orders} ta zakaz</b>\n"
    else:
        text += "<i>Hozircha haftalik reyting shakllanmoqda...</i>\n"
    text += "\n🔥 <i>Ko'proq buyurtma bajaring va haftalik maxsus bonuslarga ega bo'ling!</i>"
    await message.answer(text, reply_markup=user_main_kb(lang, uid))


@router.message(F.text.in_(["📢 Yangiliklar / Guruh", "📢 Новости / Группа"]))
async def group_handler(message: Message) -> None:
    lang = await get_lang(message.from_user.id)
    btn_txt = "💬 Haydovchilar guruhiga qo'shilish" if lang == "uz" else "💬 Вступить в группу водителей"
    await message.answer(f"📢 <b>{BOT_NAME} Rasmiy Guruhimiz:</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text=btn_txt, url=DRIVER_GROUP_LINK)]]))


@router.message(F.text.in_(["🆘 Yordam / SOS", "🆘 Помощь / SOS"]), StateFilter("*"))
async def sos_handler(message: Message, state: FSMContext) -> None:
    await state.clear()
    lang = await get_lang(message.from_user.id)
    await message.answer(t(lang, "sos_title"), reply_markup=sos_menu_kb(lang))


@router.callback_query(F.data == "sos:loc")
async def sos_location_flow(callback: CallbackQuery, state: FSMContext) -> None:
    lang = await get_lang(callback.from_user.id)
    await state.set_state(SOSStates.waiting_for_location)
    try:
        await callback.message.delete()
    except Exception:
        pass
    await callback.message.answer(t(lang, "sos_ask_loc"), reply_markup=location_request_kb(lang))
    await callback.answer()


@router.message(SOSStates.waiting_for_location, F.location)
async def sos_receive_location_geo(message: Message, state: FSMContext) -> None:
    uid = message.from_user.id
    lang = await get_lang(uid)
    await state.clear()
    user = await db_get_user(uid) or {}
    lat, lon = message.location.latitude, message.location.longitude
    maps_url = f"https://maps.google.com/?q={lat},{lon}"

    u_name = user.get("full_name", "Noma'lum")
    u_pos = user.get("position", "N/A")
    u_phone = user.get("phone", "Noma'lum")
    u_model = user.get("car_model", "")
    u_num = user.get("car_number", "")

    alert = (
        f"🚨 <b>DIQQAT: SOS / LOKATSIYA!</b>\n\n"
        f"👤 {u_name} (<code>{u_pos}</code>)\n"
        f"📱 <code>{u_phone}</code>\n"
        f"🚗 {u_model} ({u_num})\n\n"
        f"📍 <a href='{maps_url}'>Google Xaritada ochish</a>"
    )
    adm_kb = InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="💬 Haydovchi bilan chat", url=f"tg://user?id={uid}")]])
    for adm in ADMIN_IDS:
        try:
            await bot.send_message(adm, alert, reply_markup=adm_kb)
            await bot.send_location(adm, latitude=lat, longitude=lon)
        except Exception:
            pass
    await message.answer(t(lang, "sos_sent"), reply_markup=user_main_kb(lang, uid))


@router.message(SOSStates.waiting_for_location, F.text)
async def sos_receive_location_text(message: Message, state: FSMContext) -> None:
    uid = message.from_user.id
    lang = await get_lang(uid)
    if message.text in CANCEL_TEXTS:
        await state.clear()
        await message.answer(t(lang, "action_cancelled"), reply_markup=user_main_kb(lang, uid))
        return
    await state.clear()
    user = await db_get_user(uid) or {}

    u_name = user.get("full_name", "Noma'lum")
    u_pos = user.get("position", "N/A")
    u_phone = user.get("phone", "Noma'lum")
    u_model = user.get("car_model", "")
    u_num = user.get("car_number", "")
    address_txt = message.text.strip()

    alert = (
        f"🚨 <b>SOS / MANZIL (MATN):</b>\n\n"
        f"👤 {u_name} (<code>{u_pos}</code>)\n"
        f"📱 <code>{u_phone}</code>\n"
        f"🚗 {u_model} ({u_num})\n\n"
        f"📍 <b>Manzil:</b>\n{address_txt}"
    )
    adm_kb = InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="💬 Haydovchi bilan chat", url=f"tg://user?id={uid}")]])
    for adm in ADMIN_IDS:
        try:
            await bot.send_message(adm, alert, reply_markup=adm_kb)
        except Exception:
            pass
    await message.answer(t(lang, "sos_sent"), reply_markup=user_main_kb(lang, uid))


@router.callback_query(F.data == "sos:msg")
async def sos_message_flow(callback: CallbackQuery, state: FSMContext) -> None:
    lang = await get_lang(callback.from_user.id)
    await state.set_state(SOSStates.waiting_for_message)
    try:
        await callback.message.delete()
    except Exception:
        pass
    await callback.message.answer(t(lang, "sos_ask_msg"), reply_markup=cancel_kb(lang))
    await callback.answer()


@router.message(SOSStates.waiting_for_message)
async def sos_receive_text_message(message: Message, state: FSMContext) -> None:
    uid = message.from_user.id
    lang = await get_lang(uid)
    if message.text in CANCEL_TEXTS:
        await state.clear()
        await message.answer(t(lang, "action_cancelled"), reply_markup=user_main_kb(lang, uid))
        return
    await state.clear()
    user = await db_get_user(uid) or {}

    u_name = user.get("full_name", "Noma'lum")
    u_pos = user.get("position", "N/A")
    u_phone = user.get("phone", "Noma'lum")
    u_model = user.get("car_model", "")
    u_num = user.get("car_number", "")
    msg_txt = message.text or "[Xabar]"

    alert = (
        f"📩 <b>HAYDOVCHIDAN MUROJAAT:</b>\n\n"
        f"👤 {u_name} (<code>{u_pos}</code>)\n"
        f"📱 <code>{u_phone}</code>\n"
        f"🚗 {u_model} ({u_num})\n\n"
        f"✍️ <b>Xabar:</b>\n{msg_txt}"
    )
    adm_kb = InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="💬 Javob yozish", url=f"tg://user?id={uid}")]])
    for adm in ADMIN_IDS:
        try:
            await bot.send_message(adm, alert, reply_markup=adm_kb)
        except Exception:
            pass
    await message.answer(t(lang, "sos_sent"), reply_markup=user_main_kb(lang, uid))


# ============================================================
# 16. ADMIN PANEL
# ============================================================

@admin_router.message(F.text.in_(["🛠 Admin Panel", "🛠 Админ Панель"]), StateFilter("*"))
async def admin_open(message: Message, state: FSMContext) -> None:
    if message.from_user.id not in ADMIN_IDS:
        return
    await state.clear()
    lang = await get_lang(message.from_user.id)
    await message.answer("🛠 <b>Admin Boshqaruv Paneli:</b>", reply_markup=admin_main_kb(lang))


@admin_router.message(F.text.in_(["📊 Statistika", "📊 Статистика"]))
async def admin_stats_handler(message: Message) -> None:
    if message.from_user.id not in ADMIN_IDS:
        return
    stats = await db_get_stats()
    tot_u = stats.get("total_users", 0)
    reg_d = stats.get("registered_drivers", 0)
    y_lnk = stats.get("yandex_linked", 0)
    
    td_w = fmt_sum(stats.get("today_withdrawn", 0))
    mn_w = fmt_sum(stats.get("month_withdrawn", 0))
    tot_w = fmt_sum(stats.get("total_withdrawn", 0))
    pend_c = stats.get("pending_count", 0)
    pend_s = fmt_sum(stats.get("pending_sum", 0))

    await message.answer(
        f"📊 <b>{BOT_NAME} — To'liq Tizim Statistikasi:</b>\n\n"
        f"👥 Jami foydalanuvchilar: <b>{tot_u} ta</b>\n"
        f"🚕 Ro'yxatdan o'tgan haydovchilar: <b>{reg_d} ta</b>\n"
        f"🔗 Yandex Pro ulangan: <b>{y_lnk} ta</b>\n"
        f"➖➖➖➖➖➖➖➖➖➖\n"
        f"📅 <b>Bugun yechilgan summa:</b> <b>{td_w} so'm</b>\n"
        f"🗓 <b>Shu oyda yechilgan:</b> <b>{mn_w} so'm</b>\n"
        f"💸 <b>Jami (Barcha davrda):</b> <b>{tot_w} so'm</b>\n\n"
        f"⏳ <b>Kutilayotgan arizalar:</b> <b>{pend_c} ta</b> ({pend_s} so'm)"
    )


@admin_router.message(F.text.in_(["📥 Excel Hisobot", "📥 Excel Отчет"]))
async def admin_export_excel(message: Message) -> None:
    if message.from_user.id not in ADMIN_IDS:
        return
    status_msg = await message.answer("⏳ <i>Excel hisoboti tayyorlanmoqda...</i>")
    try:
        excel_bytes = await generate_monthly_excel_report()
        now = datetime.now(TASHKENT_TZ)
        month_name = UZ_MONTHS.get(now.month, "Oy")
        now_str = now.strftime("%Y_%m_%d_%H%M")
        
        for adm in ADMIN_IDS:
            try:
                file = BufferedInputFile(excel_bytes, filename=f"Lochin_Taxi_Hisobot_{now_str}.xlsx")
                await bot.send_document(
                    chat_id=adm,
                    document=file,
                    caption=f"📊 <b>{now.year}-yil {month_name} oyi Lochin Taxi hisoboti!</b>\n<i>(Kartalar xavfsiz maskalangan)</i>"
                )
            except Exception:
                pass

        try:
            await status_msg.delete()
        except Exception:
            pass
    except Exception as e:
        logger.error(f"Excel xatosi: {e}")
        await status_msg.edit_text("❌ Excel yaratishda xatolik yuz berdi.")


@admin_router.message(F.text.in_(["🔄 Yandex Sinxronlash", "🔄 Синхронизация Яндекс"]))
async def admin_sync_all_drivers(message: Message) -> None:
    if message.from_user.id not in ADMIN_IDS:
        return
    status_msg = await message.answer("⏳ <i>Yandex kabinetdagi barcha haydovchilar tekshirilmoqda...</i>")
    
    drivers, err_msg = await yandex_api.get_all_drivers(force_refresh=True)
    if not drivers:
        err_detail = err_msg if err_msg else "Noma'lum xatolik"
        await status_msg.edit_text(
            f"❌ <b>Yandex API dan ma'lumot olib bo'lmadi!</b>\n\n"
            f"🔍 <b>Xatolik sababi:</b>\n<code>{err_detail}</code>\n\n"
            f"📌 <i>Tekshiring:</i>\n"
            f"• <code>YANDEX_API_KEY</code>\n"
            f"• <code>YANDEX_CLIENT_ID</code>\n"
            f"• <code>YANDEX_PARK_ID</code>"
        )
        return

    updated_count = 0
    now = tashkent_now_iso()
    for raw_drv in drivers:
        norm = yandex_api._normalize(raw_drv)
        y_id = norm["id"]
        bal = int(norm["balance"])
        car_num = norm["car_number"]
        if not y_id:
            continue
            
        if db_pool:
            async with db_pool.acquire() as conn:
                res = await conn.execute(
                    """UPDATE users SET 
                        full_name=$1, car_model=$2, car_number=$3, 
                        balance=$4, updated_at=$5 
                    WHERE yandex_driver_id=$6 AND (balance != $4 OR car_number != $3)""",
                    norm["full_name"], norm["car_model"], car_num, bal, now, y_id,
                )
                if res != "UPDATE 0":
                    updated_count += 1
        else:
            conn = sqlite3.connect(DB_PATH, timeout=10)
            with conn:
                cur = conn.execute(
                    """UPDATE users SET 
                        full_name=?, car_model=?, car_number=?, 
                        balance=?, updated_at=? 
                    WHERE yandex_driver_id=? AND (balance != ? OR car_number != ?)""",
                    (norm["full_name"], norm["car_model"], car_num, bal, now, y_id, bal, car_num),
                )
                if cur.rowcount > 0:
                    updated_count += 1
            conn.close()

    tot_drv = len(drivers)
    await status_msg.edit_text(
        f"✅ <b>Yandex sinxronlash muvaffaqiyatli yakunlandi!</b>\n\n"
        f"🚕 Jami Yandex haydovchilari: <b>{tot_drv} ta</b>\n"
        f"🔄 Yangilanganlar: <b>{updated_count} ta</b>"
    )


@admin_router.message(F.text.in_(["👥 Haydovchilar", "👥 Водители"]))
async def admin_list_drivers(message: Message) -> None:
    if message.from_user.id not in ADMIN_IDS:
        return
    drivers = await db_get_all_registered_drivers()
    if not drivers:
        await message.answer("Hozircha ro'yxatdan o'tgan haydovchilar yo'q.")
        return

    text = f"👥 <b>Barcha Haydovchilar Ro'yxati (Jami: {len(drivers)} ta):</b>\n\n"
    for idx, drv in enumerate(drivers, 1):
        d_pos = drv.get("position", "N/A")
        d_name = drv.get("full_name", "Haydovchi")
        d_phone = drv.get("phone", "")
        d_model = drv.get("car_model", "")
        d_num = drv.get("car_number", "")
        d_bal = fmt_sum(drv.get("balance", 0))
        d_card_mask = mask_card(drv.get("card_number", ""))
        item = (
            f"<b>{idx}.</b> 🆔 <code>{d_pos}</code> — <b>{d_name}</b>\n"
            f"   📱 {d_phone} | 🚗 {d_model} ({d_num})\n"
            f"   💳 {d_card_mask} | 💰 Balans: <b>{d_bal} so'm</b>\n---------------------------\n"
        )
        if len(text) + len(item) > 4000:
            await message.answer(text)
            text = ""
        text += item

    if text:
        await message.answer(text)


@admin_router.message(F.text.in_(["🗑 Haydovchini o'chirish", "🗑 Удалить водителя"]), StateFilter("*"))
async def admin_delete_driver_prompt(message: Message, state: FSMContext) -> None:
    if message.from_user.id not in ADMIN_IDS:
        return
    await state.set_state(AdminDeleteDriverStates.waiting_for_query)
    lang = await get_lang(message.from_user.id)
    await message.answer(
        "🗑 <b>Haydovchini o'chirish bo'limi:</b>\n\n"
        "O'chirmoqchi bo'lgan haydovchining <b>POSITION ID</b>sini (masalan: <code>LCH-1416</code>) "
        "yoki <b>Telefon raqami</b>ni yuboring:\n\n"
        "<i>Bekor qilish uchun '❌ Bekor qilish' tugmasini bosing.</i>",
        reply_markup=cancel_kb(lang)
    )


@admin_router.message(AdminDeleteDriverStates.waiting_for_query)
async def admin_delete_driver_find(message: Message, state: FSMContext) -> None:
    if message.from_user.id not in ADMIN_IDS:
        return
    lang = await get_lang(message.from_user.id)
    if message.text in CANCEL_TEXTS:
        await state.clear()
        await message.answer("❌ Amaliyot bekor qilindi.", reply_markup=admin_main_kb(lang))
        return

    query = (message.text or "").strip()
    driver = await db_find_driver_by_query(query)
    if not driver:
        await message.answer(
            f"❌ <b>'{query}' bo'yicha haydovchi topilmadi!</b>\nIltimos, POSITION ID yoki telefonni to'g'ri kiriting:",
            reply_markup=cancel_kb(lang)
        )
        return

    await state.clear()
    d_id = driver["id"]
    d_pos = driver.get("position", "N/A")
    d_name = driver.get("full_name", "Haydovchi")
    d_phone = driver.get("phone", "")
    d_car = f"{driver.get('car_model','')} ({driver.get('car_number','')})"
    d_bal = fmt_sum(driver.get("balance", 0))

    info_txt = (
        f"⚠️ <b>Haqiqatdan ham ushbu haydovchini o'chirmoqchimisiz?</b>\n\n"
        f"🆔 POSITION: <code>{d_pos}</code>\n"
        f"👤 Ism: <b>{d_name}</b>\n"
        f"📱 Telefon: <code>{d_phone}</code>\n"
        f"🚗 Mashina: <b>{d_car}</b>\n"
        f"💰 Joriy balans: <b>{d_bal} so'm</b>\n\n"
        f"<i>Diqqat: Haydovchi va uning arizalari bazadan butunlay o'chiriladi.</i>"
    )

    confirm_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="🗑 Ha, o'chirish", callback_data=f"del_confirm:{d_id}"),
        InlineKeyboardButton(text="❌ Bekor qilish", callback_data="del_cancel"),
    ]])
    await message.answer(info_txt, reply_markup=confirm_kb)


@admin_router.callback_query(F.data.startswith("del_confirm:"))
async def admin_delete_driver_confirm(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Ruxsat yo'q!", show_alert=True)
        return

    user_id = int(callback.data.split(":")[1])
    success = await db_delete_user_by_id(user_id)
    if success:
        await callback.message.edit_text("✅ <b>Haydovchi bazadan muvaffaqiyatli o'chirildi!</b>")
    else:
        await callback.message.edit_text("❌ O'chirishda xatolik yuz berdi.")
    await callback.answer("Bajarildi!")


@admin_router.callback_query(F.data == "del_cancel")
async def admin_delete_driver_cancel(callback: CallbackQuery):
    await callback.message.edit_text("❌ Haydovchini o'chirish bekor qilindi.")
    await callback.answer()


@admin_router.message(F.text.in_(["📢 Xabar tarqatish", "📢 Рассылка"]))
async def admin_broadcast_prompt(message: Message, state: FSMContext) -> None:
    if message.from_user.id not in ADMIN_IDS:
        return
    await state.set_state(AdminBroadcastStates.waiting_for_message)
    lang = await get_lang(message.from_user.id)
    await message.answer("📢 <b>Barcha haydovchilarga yubormoqchi bo'lgan xabaringizni yozing:</b>\n\n<i>Bekor qilish: '❌ Bekor qilish'</i>", reply_markup=cancel_kb(lang))


@admin_router.message(AdminBroadcastStates.waiting_for_message)
async def admin_broadcast_send(message: Message, state: FSMContext) -> None:
    if message.from_user.id not in ADMIN_IDS:
        return
    lang = await get_lang(message.from_user.id)
    if message.text in CANCEL_TEXTS:
        await state.clear()
        await message.answer("❌ Xabar tarqatish bekor qilindi.", reply_markup=admin_main_kb(lang))
        return
    await state.clear()
    users = await db_get_all_users()
    status_msg = await message.answer("⏳ <i>Xabar barcha haydovchilarga yuborilmoqda...</i>")
    sent = fail = 0
    for u in users:
        tg_id = u.get("telegram_id")
        if tg_id and tg_id > 0:
            try:
                await bot.copy_message(chat_id=tg_id, from_chat_id=message.chat.id, message_id=message.message_id)
                sent += 1
                await asyncio.sleep(0.08)
            except Exception:
                fail += 1
    await status_msg.edit_text(f"📢 <b>Xabar tarqatish yakunlandi!</b>\n\n✅ Yetkazildi: <b>{sent} ta</b>\n❌ Yetib bormadi: <b>{fail} ta</b>")
    await message.answer("🛠 <b>Admin Paneli:</b>", reply_markup=admin_main_kb(lang))


@admin_router.message(F.text.in_(["🚫 Nofaollar", "🚫 Неактивные"]))
async def admin_inactive_drivers(message: Message) -> None:
    if message.from_user.id not in ADMIN_IDS:
        return
    ten_days_ago = (datetime.now(TASHKENT_TZ) - timedelta(days=10)).isoformat()
    if db_pool:
        async with db_pool.acquire() as conn:
            rows = await conn.fetch(
                "SELECT position, full_name, phone, car_model, car_number, last_activity FROM users WHERE is_registered=1 AND (last_activity < $1 OR is_blocked=1) ORDER BY id DESC LIMIT 20",
                ten_days_ago
            )
            inactive = [dict(r) for r in rows]
    else:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.row_factory = sqlite3.Row
        inactive = [dict(r) for r in conn.execute(
            "SELECT position, full_name, phone, car_model, car_number, last_activity FROM users WHERE is_registered=1 AND (last_activity < ? OR is_blocked=1) ORDER BY id DESC LIMIT 20",
            (ten_days_ago,)
        ).fetchall()]
        conn.close()

    if not inactive:
        await message.answer("✅ Barcha haydovchilar faol!")
        return
    text = f"🚫 <b>10+ kundan beri faol bo'lmagan ({len(inactive)} ta):</b>\n\n"
    for drv in inactive:
        d_name = drv.get("full_name", "Noma'lum")
        d_phone = drv.get("phone", "")
        d_model = drv.get("car_model", "")
        d_num = drv.get("car_number", "")
        d_act = str(drv.get("last_activity", "Noma'lum"))[:10]
        text += f"👤 <b>{d_name}</b> | 📱 {d_phone}\n🚗 {d_model} ({d_num})\n📅 {d_act}\n---------------------------\n"
    await message.answer(text)


@admin_router.message(F.text.in_(["⬅️ Asosiy menyu", "⬅️ Главное меню"]), StateFilter("*"))
async def back_to_user_menu(message: Message, state: FSMContext) -> None:
    await state.clear()
    uid = message.from_user.id
    lang = await get_lang(uid)
    await message.answer("Asosiy menyu:", reply_markup=user_main_kb(lang, uid))


# ============================================================
# 17. AVTOMATIK SCHEDULERLAR: ERTALABKI ESLATMA, FON SINXRON VA OYLIK
# ============================================================

async def daily_morning_reminder():
    """Har kuni ertalab soat 08:00 da barcha ro'yxatdan o'tgan haydovchilarga eslatma yuboradi."""
    last_sent_day = -1
    while True:
        try:
            now = datetime.now(TASHKENT_TZ)
            if now.hour == 8 and now.minute == 0 and now.day != last_sent_day:
                last_sent_day = now.day
                drivers = await db_get_all_registered_drivers()
                text = (
                    "🕌 <b>Assalomu alaykum hurmatli haydovchilar!</b>\n\n"
                    f"🚕 <b>{BOT_NAME}</b> jamoasi eslatadi:\n\n"
                    "⚠️ Yo'l qoidalariga qat'iy amal qiling!\n"
                    "🤝 Mijozni manzilga yetkazish davomida xushmuomala bo'ling.\n"
                    "🧼 Avtomobilingiz tozaligiga e'tibor bering.\n\n"
                    "✨ <i>Barchangizga xayrli kun va barakali daromad tilaymiz!</i>"
                )
                for d in drivers:
                    tg_id = d.get("telegram_id")
                    if tg_id and tg_id > 0:
                        try:
                            await bot.send_message(tg_id, text)
                            await asyncio.sleep(0.08)
                        except Exception:
                            pass
                logger.info(f"Ertalabki eslatma {len(drivers)} ta haydovchiga yuborildi.")
        except Exception as e:
            logger.error(f"daily_morning_reminder xatosi: {e}")
        await asyncio.sleep(30)


async def yandex_auto_sync_scheduler():
    while True:
        try:
            await asyncio.sleep(1200)
            drivers, _ = await yandex_api.get_all_drivers(force_refresh=True)
            if drivers:
                now = tashkent_now_iso()
                for raw_drv in drivers:
                    norm = yandex_api._normalize(raw_drv)
                    y_id = norm["id"]
                    bal = int(norm["balance"])
                    car_num = norm["car_number"]
                    if not y_id:
                        continue

                    if db_pool:
                        async with db_pool.acquire() as conn:
                            await conn.execute(
                                """UPDATE users SET 
                                    full_name=$1, car_model=$2, car_number=$3, 
                                    balance=$4, updated_at=$5 
                                WHERE yandex_driver_id=$6 AND (balance != $4 OR car_number != $3)""",
                                norm["full_name"], norm["car_model"], car_num, bal, now, y_id,
                            )
                    else:
                        conn = sqlite3.connect(DB_PATH, timeout=10)
                        with conn:
                            conn.execute(
                                """UPDATE users SET 
                                    full_name=?, car_model=?, car_number=?, 
                                    balance=?, updated_at=? 
                                WHERE yandex_driver_id=? AND (balance != ? OR car_number != ?)""",
                                (norm["full_name"], norm["car_model"], car_num, bal, now, y_id, bal, car_num),
                            )
                        conn.close()
                logger.info(f"Fon sinxronizatsiyasi: {len(drivers)} ta Yandex haydovchi tekshirildi.")
        except Exception as e:
            logger.error(f"Avtomatik fon sinxronlash xatosi: {e}")


async def monthly_report_scheduler():
    last_report_month = -1
    while True:
        try:
            now = datetime.now(TASHKENT_TZ)
            if now.day == 1 and now.hour == 9 and now.month != last_report_month:
                last_report_month = now.month
                excel_bytes = await generate_monthly_excel_report()
                month_name = UZ_MONTHS.get(now.month, "Oy")
                filename = f"Lochin_Taxi_{now.year}_{month_name}.xlsx"
                
                for adm in ADMIN_IDS:
                    try:
                        file = BufferedInputFile(excel_bytes, filename=filename)
                        await bot.send_document(
                            chat_id=adm,
                            document=file,
                            caption=f"🗓 <b>{now.year}-yil {month_name} Oylik Hisoboti!</b>\n\nBarcha haydovchilar statistikasi va umumiy aylanma."
                        )
                    except Exception:
                        pass
        except Exception as e:
            logger.error(f"Oylik hisobot scheduler xatosi: {e}")
        await asyncio.sleep(3600)


# ============================================================
# 18. WEB SERVER (RENDER / DOCKER HEALTH CHECKS)
# ============================================================

routes = web.RouteTableDef()


@routes.get("/")
@routes.get("/health")
async def health_check(request: web.Request) -> web.Response:
    return web.Response(text="LOCHIN TAXI ENTERPRISE 24/7 IS RUNNING ✅", status=200)


async def start_web_server():
    app = web.Application()
    app.add_routes(routes)
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", PORT)
    await site.start()
    logger.info(f"Web server 0.0.0.0:{PORT} portida muvaffaqiyatli ishga tushdi.")


# ============================================================
# 19. MAIN ASYNC RUNNER
# ============================================================

async def main() -> None:
    logger.info("Lochin Taxi Bot ishga tushirilmoqda...")
    await init_database()
    
    dp.include_router(admin_router)
    dp.include_router(router)
    
    await start_web_server()
    
    # Avtomatik fon jarayonlari (Sinxronizatsiya, Ertalabki eslatma, Oylik hisobot)
    asyncio.create_task(yandex_auto_sync_scheduler())
    asyncio.create_task(daily_morning_reminder())
    asyncio.create_task(monthly_report_scheduler())
    
    await bot.delete_webhook(drop_pending_updates=True)
    logger.info(f"{BOT_NAME} to'liq tayyor va yangilanishlarni kutmoqda!")
    
    try:
        await dp.start_polling(bot)
    finally:
        await yandex_api.close()
        if db_pool:
            await db_pool.close()


if __name__ == "__main__":
    asyncio.run(main())
